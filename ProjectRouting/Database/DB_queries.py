"""Helper funtions for routing table databse lookups/queries"""

import sqlite3
from openpyxl import Workbook
import os

mydb = sqlite3.connect("Routing")
cursor = mydb.cursor()
cursor_2 = mydb.cursor()
route_tables = []


# Begin DB Table information-----------------------------------------------


def get_tables_names() -> None:
    """Get database table names"""

    try:
        for row in cursor.execute('SELECT name FROM sqlite_master WHERE type=\'table\''):
            route_tables.append(row[0])
    except sqlite3.OperationalError:
        pass


def get_db_tables_with_data() -> list:
    """Gets database tables. If table is empty pass"""

    full_dbs = []
    get_tables = cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    for table in get_tables:
        check_table_rows = cursor_2.execute(F'SELECT count(*) FROM {table[0]}')
        for row in check_table_rows:
            if row[0] == 0:
                pass
            else:
                full_dbs.append(table[0])

    return full_dbs


# End DB Table information-----------------------------------------------

# Begin DB Quiries-------------------------------------------------------


def query_db_nexus(vdc, vrf, query, index) -> None:
    """Find routes based off query, can be full route with prefix, no mask, or just octets (192.168.)"""

    vdc_query = cursor.execute('SELECT * FROM Routing_Nexus WHERE vdc=?', (vdc,))
    matched_query = 0
    for row in vdc_query:
        if vrf in row[1] and query in row[index] and "," in row[index]:
            print(
                f"\nVDC: {row[0]}\nVRF: {row[1]}\nPrefix: {row[2]}\nProtocol: {row[3]}\nAdmin-Distance: {row[4]}\n"
                f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[7]}\nTag: {row[8]}\nAge: {row[9]}")

            matched_query += 1

        elif vrf == row[1] and query == row[index]:
            print(
                f"\nVDC: {row[0]}\nVRF: {row[1]}\nPrefix: {row[2]}\nProtocol: {row[3]}\nAdmin-Distance: {row[4]}\n"
                f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[7]}\nTag: {row[8]}\nAge: {row[9]}")

            matched_query += 1

    print(f"\nTotal Routes: {matched_query}")


def query_db_asa(query, index) -> None:
    """Find databse entries with arbitrary routing attributes. Checks for single hope metric or multi path
    using \',\' between metrics"""

    context_query = cursor.execute('SELECT * FROM Routing_ASA WHERE context=?', ("None",))
    matched_query = 0
    for row in context_query:
        if query in row[index] and "," in row[index]:
            print(f"\nContext: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
                  f"Hop(s): {row[4]}\nOut-Interface(s): {row[5]}\nMetric(s): {row[6]}\nTag: {row[7]}\nAge: {row[8]}")

            matched_query += 1

        elif query == row[index]:
            print(f"\nContext: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
                  f"Hop(s): {row[4]}\nOut-Interface(s): {row[5]}\nMetric(s): {row[6]}\nTag: {row[7]}\nAge: {row[8]}")

            matched_query += 1

    print(f"\nTotal Routes: {matched_query}")


def query_db_ios(vrf, query, index) -> None:
    """Find databse entries with arbitrary routing attributes"""

    vrf_query = cursor.execute('SELECT * FROM Routing_IOS_XE WHERE vrf=?', (vrf,))
    matched_query = 0
    for row in vrf_query:
        if query in row[index] and "," in row[index]:
            print(f"\nVRF: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
                  f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[4]}\nTag: {row[7]}\nAge: {row[8]}")

            matched_query += 1

        elif query == row[index]:
            print(f"\nVRF: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
                  f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[4]}\nTag: {row[7]}\nAge: {row[8]}")

            matched_query += 1

    print(f"\nTotal Routes: {matched_query}")


def query_db_ios_routes(vrf, query, index) -> None:
    """Find routes based off query, can be full route with prefix, no mask, or just octets (192.168.)"""

    vrf_query = cursor.execute('SELECT * FROM Routing_IOS_XE WHERE vrf=?', (vrf,))
    matched_query = 0
    for row in vrf_query:
        if query in row[index]:
            print(f"\nVRF: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
                  f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[4]}\nTag: {row[7]}\nAge: {row[8]}\n")

            matched_query += 1

    print(f"\nTotal Routes: {matched_query}")


def query_db_nexus_routes(vdc, vrf, query, index) -> None:
    """Find routes based off query, can be full route with prefix, no mask, or just octets (192.168.)"""

    vdc_query = cursor.execute('SELECT * FROM Routing_Nexus WHERE vdc=?', (vdc,))
    matched_query = 0
    for row in vdc_query:
        if vrf == row[1] and query in row[index]:
            print(f"VDC: {row[0]}\nVRF: {row[1]}\nPrefix: {row[2]}\nProtocol: {row[3]}\nAdmin-Distance: {row[4]}\n"
                  f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[7]}\nTag: {row[8]}\nAge: {row[9]}\n")

            matched_query += 1

    print(f"\nTotal Routes: {matched_query}")


def query_db_asa_routes(query, index) -> None:
    """Find routes based off query, can be full route with prefix, no mask, or just octets (192.168.)"""

    context_query = cursor.execute('SELECT * FROM Routing_ASA WHERE context=?', ("None",))

    matched_query = 0
    for row in context_query:
        if query in row[index]:
            print(f"\nContext: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
                  f"Hop(s): {row[4]}\nOut-Interface(s): {row[5]}\nMetric(s): {row[6]}\nTag: {row[7]}\nAge: {row[8]}")

            matched_query += 1

    print(f"\nTotal Routes: {matched_query}")


def view_routes_asa(table) -> None:
    """View all database entries, no filters"""

    print(f"\nRouting Table: {table}\n__________________\n")
    query = cursor.execute('SELECT * FROM Routing_ASA')
    matched_query = 0
    for row in query:
        print(f"\nContext: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
              f"Hop(s): {row[4]}\nOut-Interface(s): {row[5]}\nMetric(s): {row[6]}\nTag: {row[7]}\nAge: {row[8]}\n")
        matched_query += 1

    print(f"Total Routes: {matched_query}")


def view_routes_ios(table) -> None:
    """View all database entries, no filters"""

    print(f"\nRouting Table: {table}\n__________________\n")
    query = cursor.execute('SELECT * FROM Routing_IOS_XE')
    matched_query = 0
    for row in query:
        print(f"\nVRF: {row[0]}\nPrefix: {row[1]}\nProtocol: {row[2]}\nAdmin-Distance: {row[3]}\n"
              f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[4]}\nTag: {row[7]}\nAge: {row[8]}\n")
        matched_query += 1

    print(f"Total Routes: {matched_query}")


def view_routes_nexus(table) -> None:
    """View all database entries, no filters"""

    print(f"\nRouting Table: {table}\n__________________\n")
    query = cursor.execute('SELECT * FROM Routing_Nexus')
    matched_query = 0
    for row in query:
        print(f"VDC: {row[0]}\nVRF: {row[1]}\nPrefix: {row[2]}\nProtocol: {row[3]}\nAdmin-Distance: {row[4]}\n"
              f"Hop(s): {row[5]}\nOut-Interface(s): {row[6]}\nMetric(s): {row[7]}\nTag: {row[8]}\nAge: {row[9]}\n")
        matched_query += 1

    print(f"Total Routes: {matched_query}")


def export_excel(table):
    """Reads database and writes to excel file"""

    routes = Workbook()
    dir_path = os.path.dirname(os.path.realpath(__file__))
    routes_file = dir_path + '\\RoutingTable.xlsx'
    active_sheet = routes.active
    get_tables_names()

    query = cursor.execute('SELECT * FROM %s' % table)
    excel_row = 1
    for row in query:
        active_sheet.cell(row=excel_row, column=1, value=row[0])
        active_sheet.cell(row=excel_row, column=2, value=row[1])
        active_sheet.cell(row=excel_row, column=3, value=row[2])
        active_sheet.cell(row=excel_row, column=4, value=row[3])
        active_sheet.cell(row=excel_row, column=5, value=row[4])
        active_sheet.cell(row=excel_row, column=6, value=row[5])
        active_sheet.cell(row=excel_row, column=7, value=row[6])
        active_sheet.cell(row=excel_row, column=8, value=row[7])
        active_sheet.cell(row=excel_row, column=9, value=row[8])
        if table == "Routing_Nexus":
            active_sheet.cell(row=excel_row, column=10, value=row[9])
        excel_row += 1
    check_permission(routes, routes_file)


def check_permission(workbook, filepath):
    try:
        workbook.save(filename=filepath)
    except OSError:
        print(f"Unable to save file, check permission for {filepath}")
        input("Press Enter to close")


# End DB Quiries-------------------------------------------------------

# Begin query Helpers. Prequery attributes----------------------------


def get_routing_interfaces(table) -> dict:
    """Gets routing interfaces from table"""

    interfaces = {}
    get_interfaces = cursor.execute(f'SELECT interfaces FROM {table}')
    for row in get_interfaces:
        if row[0].rfind(", ") != -1:
            split_interfaces = row[0].split(", ")
            for i in split_interfaces:
                interfaces[i[0]] = None
        else:
            interfaces[row] = None

    return interfaces


def print_routing_interfaces(table) -> None:
    """Gets routing interfaces from table, prints"""

    interfaces = get_routing_interfaces(table=table)

    print("\nRouting Interfaces ---------\n")
    if len(interfaces) == 0:
        pass
    else:
        for k in interfaces.keys():
            if len(k[0]) < 2 or "None" in k[0]:
                pass
            else:
                print("+ " + k[0])
        print("\n")


def get_protocols(table) -> dict:
    """Gets route protocol with types from the the routing table"""

    protocol = {}
    get_protocol = cursor.execute(f'SELECT interfaces FROM {table}')
    for i in get_protocol:
        protocol[i[0][0]] = "None"

    return protocol


def get_vrfs(table) -> None:
    """Gets VRFs from the device. Databse default is "global"""

    vrf = {}
    vrfs = cursor.execute(f'SELECT vrf FROM {table}')

    for i in vrfs:
        vrf[i] = None

    print("\nVRFs ---------\n")
    if len(vrf) == 0:
        print("\nPress enter to use global")
    else:
        for k in vrf.keys():
            print(f"+ {k[0]}")
        print("\n")


def get_vdcs():
    """Gets VDCs from the device table"""

    vdc = {}
    vdcs = cursor.execute('SELECT vdc FROM Routing_Nexus')
    for i in vdcs:
        vdc[i] = None

    print("\nVDCs ---------\n")
    if len(vdc) == 0:
        pass
    else:
        for k in vdc.keys():
            print(f"+ {k[0]}")
        print("\n")


def get_admin_disatnces(table) -> None:
    """Gets administrative distance from the routing table"""

    ad = {}
    get_ads = cursor.execute(f'SELECT admin_distance FROM {table}')
    for i in get_ads:
        ad[i] = None

    print("\nAdmin Distances ---------\n")
    if len(ad) == 0:
        pass
    else:
        for k in ad.keys():
            print(f"+ {k[0]}")
        print("\n")


def get_tags(table) -> None:
    """Gets tags from the the routing table"""

    tag = {}
    tags = cursor.execute(f'SELECT tag FROM {table}')
    for i in tags:
        tag[i] = None

    print("\nRoute Tags ---------\n")
    if len(tag) == 0:
        pass
    else:
        for k in tag.keys():
            print(f"+ {k[0]}")
        print("\n")


def print_protocols(table) -> None:
    """Gets route protocol with types from the the routing table"""

    protocol = {}
    get_protocol = cursor.execute(f'SELECT protocol FROM {table}')
    for i in get_protocol:
        protocol[i] = None

    print("\nProtocols ---------\n")
    if len(protocol) == 0:
        pass
    else:
        for k in protocol.keys():
            if k[0].rfind("None") != -1:
                pass
            else:
                print(f"+ {k[0]}")
        print("\n")


# End query Helpers. Prequery attributes----------------------------

# Begin query builders----------------------------------------------


def search_db_ios(vrf=None, prefix=None, protocol=None, metric=None, ad=None, tag=None, interface=None) -> None:
    """Find databse entries by artbitrary attribute using **attributes (kwargs)
                            vrf, admin-distance, metric, prefix, next-hop, tag"""

    if vrf is None or vrf == "":
        vrf = "global"
    else:
        pass

    if protocol is not None:
        query_db_ios(vrf=vrf, query=protocol, index=2)
    if prefix is not None:
        query_db_ios_routes(vrf=vrf, query=prefix, index=1)
    if metric is not None:
        query_db_ios(vrf=vrf, query=metric, index=4)
    if ad is not None:
        query_db_ios(vrf=vrf, query=ad, index=3)
    if tag is not None:
        query_db_ios(vrf=vrf, query=tag, index=7)
    if interface is not None:
        query_db_ios(vrf=vrf, query=interface, index=6)


def search_db_nexus(vdc=None, vrf=None, prefix=None, protocol=None, metric=None, ad=None, tag=None,
                    interface=None) -> None:
    """Find databse entries by artbitrary attribute using **attributes (kwargs)
                            vrf, admin-distance, metric, prefix, next-hop, tag"""

    if vrf is None or vrf == "":
        vrf = "default"
    else:
        pass

    if protocol is not None:
        query_db_nexus(vdc=vdc, vrf=vrf, query=protocol, index=3)
    if prefix is not None:
        query_db_nexus_routes(vdc=vdc, vrf=vrf, query=prefix, index=2)
    if metric is not None:
        query_db_nexus(vdc=vdc, vrf=vrf, query=metric, index=7)
    if ad is not None:
        query_db_nexus(vdc=vdc, vrf=vrf, query=ad, index=4)
    if tag is not None:
        query_db_nexus(vdc=vdc, vrf=vrf, query=tag, index=8)
    if interface is not None:
        query_db_nexus(vdc=vdc, vrf=vrf, query=interface, index=6)


def search_db_asa(prefix=None, protocol=None, metric=None, ad=None, tag=None, interface=None) -> None:
    """Find databse entries by artbitrary attribute using **attributes (kwargs)
                            vrf, admin-distance, metric, prefix, next-hop, tag"""

    if protocol is not None:
        query_db_asa(query=protocol, index=2)
    if prefix is not None:
        query_db_asa_routes(query=prefix, index=1)
    if metric is not None:
        query_db_asa(query=metric, index=6)
    if ad is not None:
        query_db_asa(query=ad, index=3)
    if tag is not None:
        query_db_asa(query=tag, index=7)
    if interface is not None:
        query_db_asa(query=interface, index=5)

# End query builders----------------------------------------------
