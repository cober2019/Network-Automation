try:
    import sqlite3
except ImportError:
    print("Module SQLITE3 not available.")
try:
    from ncclient import manager
except ImportError:
    print("Module NCC Client not available.")
try:
    import ncclient
except ImportError:
    print("Module NCC Client not available.")
try:
    from socket import gaierror
except ImportError:
    print("Module socket not available.")
try:
    from datetime import date
except ImportError:
    print("Module time not available.")
try:
    import time
except ImportError:
    print("Module time not available.")
try:
    from openpyxl import load_workbook
except ImportError:
    print("Module openpyxl not available.")
try:
    import xmltodict
except ImportError:
    print("Module xmltodict not available.")

mydb = sqlite3.connect("Automation")
c = mydb.cursor()
d = mydb.cursor()
status_1 = "Pass"
status_2 = "Fail"
global_status = " "

def main():

    config_selection = ' '
    while config_selection != 'q':

        print("\n")
        print("Basic Network Programabiltiy Database")
        print("\n")
        print(" 1: Add Device")
        print(" 2: Delete Device")
        print(" 3: View Database")
        print(" 4: Search DB")
        print(" 5: Delete Database")
        print(" 6: Device Update")
        print(" 7: Capabilities")
        print(" 8: Device Import")
        print("\n")

        print("[q] (quit)")

        print("\n")
        config_selection = input("Please select an option:  ")
        print("\n")

        if config_selection == "1":
            db_check()
            manual_CreateDB()
        if config_selection == "2":
            db_check()
            delete_device()
        if config_selection == "3":
            db_check()
            view_database()
        if config_selection == "4":
            db_check()
            find_db_specific()
        if config_selection == "5":
            db_check()
            delete_database()
        if config_selection == "6":
            db_check()
            device_update()
        if config_selection == "7":
            db_check()
            get_capabilities()
        if config_selection == "8":
            db_check()
            import_devices()


def db_check():

    # auto creates db. This is to prevent any exceptions withing th program. Ran at main()

    try:
        c.execute('''CREATE TABLE Automation (Device, DevType, Location, Status, DateTime, Time, Model, Serial)''')
        mydb.commit()
    except sqlite3.OperationalError:
        pass

def manual_CreateDB():

    # manually add devices to DB, default column entries

    db_check()

    Device = input("Device: ")
    Location = input("Location: ")
    abbriviate = Location[0:3].upper()

    try:
        c.execute("INSERT INTO Automation VALUES ('%s','Protocol','%s', 'Unknown   ', '%s', '%s', '%s', '%s')"
                  % (Device, abbriviate, date.today(), time.strftime("%H:%M:%S ", "Model", "Serial")))
        mydb.commit()
    except sqlite3.OperationalError:
        pass

    cleanup_db()

def headers():

    #Any time database is viewed, filtered or unfiltered

    print("\n")
    print("{:>12} {:>24} {:>13} {:>14} {:>17} {:>16} {:>18} {:>16}".format("IP", "Protocol", "Location", "Status", "Date", "Time", "Model", "Serial"))
    print("{:>14} {:>23} {:>13} {:>14} {:>16} {:>17} {:>18} {:>15}".format("____ ", "________", "________", "______", "____", "______", "______", "______"))
    print("\n")

def view_database():

    #  view entire unfiltered database, headers() as headers

    try:
        headers()
        for row in c.execute('SELECT * FROM Automation'):
            print("{:>0} {:>15} {:>19} {:>11} {:>14} {:>21} {:>17} {:>16} {:>15}".format("  ", row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))
        print("\n")

    except sqlite3.OperationalError:
        print("\n")
        print("Database Doesn't Exist")
        print("\n")
        pass

def delete_database():
    try:
        c.execute('''DROP TABLE Automation''')
        mydb.commit()
    except sqlite3.OperationalError:
        pass

def find_db_specific():

    # Filter devices based off Location, IP, and Pass/Fail/Unknown
    # headers() function is preformated heading for DB columns

    print("\n")
    print("1. Location")
    print("2. IP")
    print("3. Pass/Fail/New")
    print("4. Serial")
    print("5. Main")
    print("\n")

    while True:

        input_selection = input("Selection: ")
        print("\n")

        if input_selection == "1":

            location = input("Location: ")
            abbreviate = location[0:3].upper()

            # Loops through rows in the DB, prints filtered results

            try:
                headers()
                for row in c.execute('SELECT * FROM Automation WHERE Location=?', (abbreviate, )):
                    print("{:>0} {:>15} {:>19} {:>11} {:>14} {:>21} {:>17} {:>16} {:>15}".format(
                    "  ", row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))
                find_db_specific()
            except sqlite3.OperationalError:
                pass

            find_db_specific()
            break

        elif input_selection == "2":

            ip = input("IP: ")

            # Loops through rows in the DB, prints filtered results

            try:
                headers()
                for row in c.execute('SELECT * FROM Automation WHERE Device=?', (ip, )):
                    print("{:>0} {:>15} {:>19} {:>11} {:>14} {:>21} {:>17} {:>16} {:>15}".format(
                    "  ", row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))

                find_db_specific()
                break

            except sqlite3.OperationalError:
                pass
            find_db_specific()

        elif input_selection == "3":

            status = input("Status: ")

            # Loops through rows in the DB, prints filtered results

            try:
                headers()
                for row in c.execute('SELECT * FROM Automation WHERE Status=?', (status, )):
                    print("{:>0} {:>15} {:>19} {:>11} {:>14} {:>21} {:>17} {:>16} {:>15}".format(
                    "  ", row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))

                find_db_specific()
                break

            except sqlite3.OperationalError:
                break

            find_db_specific()

        elif input_selection == "4":

            serial = input("Serial: ")

            # Loops through rows in the DB, prints filtered results

            try:
                headers()
                for row in c.execute('SELECT * FROM Automation WHERE Serial=?', (serial, )):
                    print("{:>0} {:>15} {:>19} {:>11} {:>14} {:>21} {:>17} {:>16} {:>15}".format(
                    "  ", row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))

                find_db_specific()
                break

            except sqlite3.OperationalError:
                break

            find_db_specific()

        elif input_selection == "5":
            main()
            break
        else:
            print("\n")
            print("Invalid Selection")
            print("\n")
            continue

def delete_device():

    # Delete single device via IP

    view_database()

    print("\n")
    device = input("Device: ", )
    print("\n")

    try:
        c.execute('DELETE FROM Automation WHERE Device=?', (device, ))
        mydb.commit()
    except sqlite3.OperationalError:
        pass

def device_update():

    # Update DB entry attributes. Just in case you need to change location or IP of a device

    column_array = ["DevType", "Location"]

    view_database()
    device = input("Device: ")

    print("\n")
    print("-Attributes-")
    print("\n")
    print("Location")
    print("Devtype")
    print("\n")

    column_Selection = input("Attribute: ")
    new_attrib = input("New Attribute: ")

    try:
        c.execute('SELECT * FROM Automation WHERE Device=?', (device,))
        if column_Selection == "Location":
            c.execute("UPDATE Automation SET Location=?, DateTime=?, Time=? WHERE Device=?",
                      (new_attrib, date.today(), time.strftime("%H:%M:%S "), device,))
            mydb.commit()

        elif column_Selection == "DevType":
            c.execute("UPDATE Automation SET DevType=?, DateTime=?, Time=? WHERE Device=?",
                      (new_attrib, date.today(), time.strftime("%H:%M:%S "), device,))
            mydb.commit()

    except sqlite3.OperationalError:
        pass

def db_update(host, status, devtype, model, serial):

    try:
        d.execute("UPDATE Automation SET DevType=?, Time=?, DateTime=?, Status=?, Model=?, Serial=? WHERE Device=?",
                  (devtype, time.strftime("%H:%M:%S "), date.today(), status, model, serial, host))
        mydb.commit()
    except sqlite3.OperationalError:
        pass

def device_connect(host):

        # NETCONF login
        # db_update function updates the DB entry based on pass/fail
        # global status variable keeps track of device connectivty, later used in get_capabilties function.
        # If status is "fail" program won't check for capabilities

        global global_status
        try:

            global m
            m = manager.connect(host=host, port=830, timeout=3, username="cisco", password="cisco",
                                            device_params={'name': 'csr'})

        except ncclient.NCClientError:
            global_status = status_2
            db_update(host, status_2, "Unknown", " Unknown", "Unknown")
            pass
        except AttributeError:
            global_status = status_2
            db_update(host, status_2, "Unknown", " Unknown", "Unknown")
            pass
        except gaierror:
            global_status = status_2
            db_update(host, status_2, "Unknown", " Unknown", "Unknown")
            db_conn.start()
            pass

def capabilities(host):

    info = """
    <filter>
      <native xmlns="http://cisco.com/ns/yang/Cisco-IOS-XE-native">
        <license/>
      </native>
    </filter>
        """
    for capability in m.server_capabilities:
        db_update(host, status_1, "NET/YANG", "Unknown", "Unknown")

    try:
        config_data = m.get(info)
        config_details = xmltodict.parse(config_data.xml)["rpc-reply"]["data"]
        vlan_details = config_details["native"]["license"]["udi"]

        model = vlan_details["pid"]
        serial = vlan_details["sn"]

        db_update(host, status_1, "NET/YANG", model, serial )

    except TypeError:
        pass

def get_capabilities():

    # Get device capabilities based off IP, Location, or Status

    view_database()
    retries = 0

    print("1. IP")
    print("2. Status")
    print("3. Location")
    print("4. Main")
    print("\n")

    input_selection = input("Selection: ")
    print("\n")

    # device_connect function is NETCONF session creation indexing the first entry in the db/row

    try:
        if input_selection == "1":

            device_ip = input("IP: ")

            for row in c.execute('SELECT * FROM Automation WHERE Device=?', (device_ip,)):

                device_connect(row[0])

            else:
                main()
    except sqlite3.OperationalError:
        pass

        # Selction 1 & 2 -Check global status created on device_connect function, if failed continue loop, if pass check for device capabilities

    try:
        if input_selection == "2":

            pass_fail = input("Pass/Fail/Unknown: ")

            for row in c.execute('SELECT * FROM Automation WHERE Status=?', (pass_fail,)):

                device_connect(row[0])
                if global_status == status_2:
                    continue
                else:
                    capabilities(row[0])

            else:
                main()
    except sqlite3.OperationalError:
        pass

    try:
        if input_selection == "3":

            location = input("Location: ")

            for row in c.execute('SELECT * FROM Automation WHERE Location=?', (location, )):

                device_connect(row[0])
                if global_status == status_2:
                    continue
                else:
                    capabilities(row[0])
            else:
                main()
    except sqlite3.OperationalError:
        pass

        if input_selection == "4":
            main()

def import_devices():

    # Imports devices from excel spreadsheet.

    ip_array = [ ]
    location_array = [ ]

    try:
        wb_file = "C:\Python\Book1.xlsx"
        workbook = load_workbook(wb_file)
        active_sheet = workbook.active
        active_sheet.protection = False

        for col in active_sheet.iter_rows(min_row=1, max_col=1, max_row=100): # Read first row/column, place into an array
            for cell in col:
                if cell.value == None:
                    continue
                else:
                    ip_array.append(cell.value)

        for row in active_sheet.iter_rows(min_row=1, min_col=2, max_row=100): # Read first row/Second column, place into an array
            for cell in row:
                if cell.value == None:
                    continue
                else:
                    location_array.append(cell.value)

        for ip, location in zip(ip_array, location_array): # Read both arrays above simaltaniuously, insert ip and location into database
            try:
                c.execute("INSERT INTO Automation VALUES ('%s','Unknown','%s', 'New', '%s', '%s', '%s', '%s')" %
                          (ip, str(location[0:3]).upper(), date.today(), time.strftime("%H:%M:%S "), "Unknown", "Unknown"))
                mydb.commit()
            except TypeError:
                pass
            except sqlite3.OperationalError:
                pass

        cleanup_db()

    except FileNotFoundError:
        print("Workbook Not Found")
        pass

def cleanup_db():

    # Cleanup any duplicate DB entries, keeping the most current (max)

    try:
        for row in c.execute('SELECT * FROM Automation'): # Cleanup any database duplicates
                c.execute('DELETE FROM Automation WHERE rowid not in (SELECT max(rowid) FROM Automation GROUP BY Device)')
                mydb.commit()
    except sqlite3.OperationalError:
        pass

if __name__ == '__main__':

    main()