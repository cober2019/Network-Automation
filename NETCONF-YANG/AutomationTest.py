
try:
    from ncclient import manager
except ImportError:
    print("Module NCC Client not available.")
try:
    import  ncclient
except ImportError:
    print("Module NCC Client not available.")
try:
    from openpyxl import load_workbook
except ImportError:
    print("Module openpyxl not available.")
try:
    from openpyxl.styles import PatternFill
except ImportError:
    print("Module openpyxl not available.")
try:
    import xml.etree.ElementTree as xml
except ImportError:
    print("Module xml.etree.ElementTree not available.")
try:
    import xml.dom.minidom as dom
except ImportError:
    print("Module xml.dom.minidom not available.")
try:
    import lxml.etree as ET
except ImportError:
    print("Module lxml.etree as ET not available.")
try:
    import xmltodict
except ImportError:
    print("Module xmltodict not available.")
try:
    import pathlib
except ImportError:
    print("Module pathlib not available.")
try:
    from socket import gaierror
except ImportError:
    print("Module socket not available.")
try:
    import ipaddress
except ImportError:
    print("Module readline not available.")
try:
    import ftplib
except ImportError:
    print("Module ipaddress not available.")
try:
    import time
except ImportError:
    print("Module time not available.")
try:
    import readline
except ImportError:
    print("Module readline not available.")
try:
    from ncclient.operations import RPCError
except ImportError:
    print("Module NCC Client not available.")
try:
    import paramiko
except ImportError:
    print("Module paramiko not available.")
try:
    import pyang
except ImportError:
    print("Module pyang not available.")
try:
    import sys
except ImportError:
    print("Module sys not available.")


def disable_paging(remote_conn):

    remote_conn.send("terminal length 0\n")
    time.sleep(1)

    output = remote_conn.recv(1000)
    return output

def paramiko_login(command):

    try:
        ip = input("Please enter a IP address: ")
        username = 'cisco'
        password = 'cisco'

        remote_conn_pre = paramiko.SSHClient()

        remote_conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        remote_conn_pre.connect(ip, username=username, password=password, look_for_keys=False, allow_agent=False)
        print("SSH connection established to %s" % ip)

        remote_conn = remote_conn_pre.invoke_shell()
        print("Interactive SSH session established")

        output = remote_conn.recv(1000)
        print(output)

        disable_paging(remote_conn)

        remote_conn.send("\n")
        remote_conn.send(command)

        remote_conn.send(command)
        time.sleep(2)

        output = remote_conn.recv(5000)
        output_str = output.decode('utf-8')
        print(output_str)


        while True:
            try:

                repeat_selection = input("Do you want to repeat command? ")

                if repeat_selection == "yes":
                    remote_conn.send(command)
                    time.sleep(2)

                    output = remote_conn.recv(5000)
                    output_str = output.decode('utf-8')
                    print(output_str)
                    continue

                elif repeat_selection == "no":
                    main()
                    break
                else:
                    print("\n")
                    print("Invalid Selection")
                    print("\n")
                    continue
            except paramiko.ssh_exception:
                print("\n")
                print("Connection Unsuccessful")
                print("\n")
                main()
    except paramiko.ssh_exception:
        print("\n")
        print("Connection Unsuccessful")
        print("\n")
        device_admin()

def permit_deny_selection(text, state):

    permit_deny = ["permit", "deny"]
    perm_deny_commands = [cmd for cmd in permit_deny if cmd.startswith(text)]

    if state < len(perm_deny_commands):
        return perm_deny_commands

def ospf_net_type_selection(text, state):

    net_types = ["point-to-point", "point-to-multipoint", "broadcast"]
    net_types_commands = [cmd for cmd in net_types if cmd.startswith(text)]

    if state < len(net_types_commands):
        return net_types_commands[state]
    else:
        return None

def qos_shape_pol_selection(text, state):

    shape_pol_actions = ["shape", "police"]
    shape_pol_commands = [cmd for cmd in shape_pol_actions if cmd.startswith(text)]

    if state < len(shape_pol_commands):
        return shape_pol_commands[state]
    else:
        return None

def tunnel_mode_selection(text, state):

    tunnel_actions = ["transport", "tunnel"]
    tunnel_action_commands = [cmd for cmd in tunnel_actions if cmd.startswith(text)]

    if state < len(tunnel_action_commands):
        return tunnel_action_commands[state]
    else:
        return None

def qos_action_selection(text, state):

    qos_actions = ["bandwidth", "priority"]
    qos_action_commands = [cmd for cmd in qos_actions if cmd.startswith(text)]

    if state < len(qos_action_commands):
        return qos_action_commands[state]
    else:
        return None

def qos_match_selection(text, state):

    match_types = ["match-any", "match-all"]
    match_types_commands = [cmd for cmd in match_types if cmd.startswith(text)]

    if state < len(match_types_commands):
        return match_types_commands[state]
    else:
        return None

def int_type_selection(text, state):

    int_types = ["0/0/0", "0/0/1", "0/0/2", "0/0/3"]
    int_type_commands = [cmd for cmd in int_types if cmd.startswith(text)]

    if state < len(int_type_commands):
        return int_type_commands[state]
    else:
        return None

def int_selection(text, state):

    int_options = ["Tunnel", "GigabitEthernet", "FastEthernet", "Loopback"]
    int_opt_commands = [cmd for cmd in int_options if cmd.startswith(text)]

    if state < len(int_opt_commands):
        return int_opt_commands[state]
    else:
        return None

def cleanup_empty_elements(root_var, file):

    tree = xml.ElementTree(element=root_var)
    tree.write(file_or_filename=file)

    root = ET.fromstring(open(file=file).read())
    for element in root.xpath("//*[not(node())]"):
        element.getparent().remove(element)

    tree = xml.ElementTree(element=root)
    tree.write(file_or_filename=file)

def ftp_files():

    # DOWNLOADS FILE FROM AN FTP SERVER AND SAVES THEM TO A LOCAL DIRECTORY

    filename = "Book1.xlsx"
    dest_file = open("C:\Python" + "\\" + filename, 'wb')
    dest_path =  "C:\Python" + "\\" + filename

    print(" 1: FTP Download")
    print(" 2: Main Menu")
    print("\n")
    print("Press CTRL+C to escape at any time")
    print("\n")

    print("\n")
    question_1 = input("Please select an option: ")
    print("\n")

    if question_1 == "1":
        print("\n")
        ftp_server = input("Please enter an FTP server: ")
        print("\n")

        ftp = ftplib.FTP(ftp_server)
        ftp.login("admin", "1234")
        ftp.cwd("/")

        try:
            ftp.retrbinary("RETR " + filename, dest_file.write, 1024)
            print("File Downloaded: Path:  " + dest_path)
            time.sleep(2)
            ftp.quit()
            main()

        except ftplib.error_perm:
            print("Please check the following - Filepath, credentials, Filename")
            ftp_files()
        except TimeoutError:
            print("Could not connect to FTP server. Please check connectivity")
            ftp_files()
        except UnicodeError:
            print("Invalid IP address. Please re-enter")
            ftp_files()

    if question_1 == "2":
        main()

    else:
        print("Invalid input")
        ftp_files()

#################################################################

def select_configuration_send():

    # CONFIGURATION FILES CAN BE VIEWED FROM THIS DIRECTORY AND SENT. THIS CAN BE USED TO SEND FILES THAT INTIALLY FAILED

    print(" 1: File Select")
    print(" 2: Main Menu")

    print("\n")
    question_1 = input("Please select an option: ")
    print("\n")

    if question_1 == "1":

        dir_files = []
        for p in pathlib.Path("C:\Python\XML_Filters\Send_Config").iterdir():
            if p.is_file():
                print(p)
                dir_files.append(p)

        file = input("Please enter a filename: ")

        try:
            config_file = open("C:\Python\XML_Filters\Send_Config" + "\\" +  file).read()
            print (dom.parseString(str(config_file)).toprettyxml())

            send_payload = m.edit_config(config_file, target="running")
            save_configuration()
            main()

        except ncclient.operations.rpc.RPCError:
            print("\n")
            print("Invalid XML Element value. Please review your configurations")
            time.sleep(2)
            print("\n")


#######################################################################################


def save_configuration():

    # SAVES RUNNING CONFIG TO STARTUP CONFIG AFTER CONFIGURATION SEND

    save_payload = """
    		<cisco-ia:save-config xmlns:cisco-ia="http://cisco.com/yang/cisco-ia"/>
            """
    try:
        response = m.dispatch(ET.fromstring(save_payload))
        print("Configuration Saved")
        main()
    except ValueError:
        print("Error")
        main()

########################################################################################

def view_config_send(file):

    # DEFINES SINGLE DEVICE CONIFGURATION OR MULTI DEVICE SEND. YOU ARE ALSO ABLE TO CANCLE A SEND AS WELL

    print("\n")
    print ("Configuration: ")
    print("\n")

    Interface_config = open(file=file).read()
    print(dom.parseString(str(Interface_config)).toprettyxml())

    while True:

        print(" 1: Send Single Device Configuration")
        print(" 2: Send Multi  Device Configuration")
        print(" 3: Cancel Configuration Send")

        print("\n")
        question_1 = input("Please select an option: ")
        print("\n")

        if question_1 == "1":
            send_single_configuration(file=file)
            break
        if question_1 == "2":
            send__multi_configuration(file=file)
            break
        if question_1 == "3":
            print("\n")
            print("Configuration Canceled")
            print("\n")
            break
        else:
            print("\n")
            print("Invalid Selection")
            print("\n")



######################################################################

def ncc_login(host, port, username, password, device_params): # Log into device via NCC client

    #NCCLIENT LOGIN
    try:
        global m
        global device
        print("\n")
        device = input("Please Enter a device IP address: ")
        print("\n")
        print("\n")
        print("Press CTRL+C to escape at any time")
        print("\n")

        try:

            global m
            m = manager.connect(host=device, port=port, username=username, password=password, device_params=device_params)

            print("\n")
            print("Device:", device, "Session ID: ", m.session_id, " Connection: ",m.connected)
            print("\n")
        except UnicodeError:
            print("Invalid IP address. Insure IP format is correct")
            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
        except ncclient.NCClientError:
            print("Connection Timeout. Please check connectivity and NETCONF configuration.")
            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
        except gaierror:
            print("Invalid IP address. Insure IP format is correct")
            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
    except KeyboardInterrupt:
        main()



################################################################################

def send__multi_configuration(file):

    #READS DEVICE FROM AND EXCEL SPREAD AND FILLS THE CELL WITH EITHER GREEN OR RED. A GREEN FILLED CELL IS A SUCCESS, RED IS A FAILURE. THESE ARE FILLED BASED ON AN EXCEPTION DURING THE JOB

    success_fill = PatternFill(start_color='00FF00',
                                                end_color='00FF00',
                                                fill_type='solid')

    fail_fill = PatternFill(start_color='FF8080',
                                                end_color='FF8080',
                                                fill_type='solid')

    status = [ ]

    wb_file = "C:\Python\Book1.xlsx"
    workbook = load_workbook(wb_file)
    active_sheet = workbook.active
    active_sheet.protection = False

    for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=10):
        for cell in row:
            if cell.value == 'Null':
                print("\n")
                print("Job complete. Please review inventory sheet for send statuses. Red=Failed, Green=Success")
                time.sleep(2)
                for item in status:
                    print(item)

                try:
                    workbook.save("C:\Python\Book1.xlsx")
                    workbook.save(wb_file)
                    main()
                except PermissionError:
                    print("Could not write to file. Please ensure the file isnt open and you have write permissions.")
                    main()
            else:

                try:
                    x = manager.connect(host=cell.value, port=830, username="cisco", password="cisco", device_params={
                        'name': 'csr'})

                    print("Device:", cell.value, "Session ID: ", x.session_id, " Connection: ", x.connected)
                    print("\n")

                    config_file = open(file=file).read()
                    send_payload = x.edit_config(config_file, target="running")

                    save_payload = """
                                                    <cisco-ia:save-config xmlns:cisco-ia="http://cisco.com/yang/cisco-ia"/>
                                              """
                    try:
                        response = x.dispatch(ET.fromstring(save_payload))
                    except ValueError:
                        print("\n")
                        print("Configuration Saved")
                        status.append(cell.value + " - Success")
                        cell.fill = success_fill
                        continue
                    except AttributeError:
                        print("\n")
                        print("Connection Unsucessful")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        pass
                    except RPCError:
                        print("\n")
                        print("Payload Error")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        pass
                except UnicodeError:
                    print("\n")
                    print("Invalid IP address. Please try again")
                    status.append(cell.value + " - Failed")
                    cell.fill = fail_fill
                    pass
                except ncclient.NCClientError:
                    print("\n")
                    print("Please review configuration.")
                    status.append(cell.value + " - Failed")
                    cell.fill = fail_fill
                    pass

##################################################################################


def send_single_configuration(file):

        # USED FOR SINGLE DEVICE CONFIGURATION. ONCE CONFIGURATION IS COMPLETE, YOU WILL BE RETURNED TO THE MENU FROM WHERE YOU CONFIGURING. THIS IS BASED OF THE FILENAME OF THE CONFIGURATION

        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        try:
            config_file = open(file=file).read()
            send_payload = m.edit_config(config_file, target="running")
            save_configuration()

            print("\n")
            print("Configuration Complete!")
            print("\n")
            main()

        except ValueError:
            # This exception can be triggered but the configuration is still saved to startup-config
            print("Configuration Saved")
            main()
        except AttributeError:
            print("Connection Unsucessful")
            main()
        except RPCError:
            print("Something went wrong")
            main()


###########################################################################

def search_credentials():

    get_file = "C:\Python\XML_Filters\Credentials_Get_Config.xml"
    search = input("Search Username: ")

    success_fill = PatternFill(start_color='00FF00',
                                                end_color='00FF00',
                                                fill_type='solid')

    fail_fill = PatternFill(start_color='FF8080',
                                                end_color='FF8080',
                                                fill_type='solid')

    status = [ ]
    key = "name"

    try:
        wb_file = "C:\Python\Book1.xlsx"
        workbook = load_workbook(wb_file)
        active_sheet = workbook.active
        active_sheet.protection = False

    except BaseException:
        print("\n")
        print("Unknown file error")
        search_configurations()
    else:
        for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=10):
            for cell in row:
                if cell.value == 'Null':
                    print("\n")
                    print("Scan complete. Please review inventory sheet for send statuses. Red=Failed Connection, Green=Success", "No File =  User not Found")
                    time.sleep(2)
                    print("\n")
                    remove_duplicates = (set(status))
                    for items in remove_duplicates:
                        print(items)

                    try:
                        workbook.save("C:\Python\Book1.xlsx")
                        workbook.save(wb_file)
                        main()
                    except PermissionError:
                        print("Could not write to file. Please ensure the file isnt open and you have write permissions.")
                        main()
                else:
                    try:
                        m = manager.connect(host=cell.value, port=830, username="cisco", password="cisco", device_params={ 'name': 'csr'})

                        print("\n")
                        print("Device:", cell.value, "Session ID: ", m.session_id, " Connection: ", m.connected)
                        print("\n")

                        credential_config = open("C:\Python\XML_Filters\Filter_Config.xml").read()
                        config_data = m.get(credential_config)

                        tree = xml.ElementTree("data")  # Writes XML file to share
                        with open(get_file, "wb") as fh:
                            tree.write(fh)

                    except ncclient.NCClientError:
                        print("\n")
                        print("Connection Unsuccessful to " +  cell.value)
                        cell.fill = fail_fill
                        status.append(cell.value + " - Failed")
                        pass

###################################################

def search_snmp():

    # Search config for a given configuration uses SEARCH functions to convert XML to dictionary.It then runs a for loop for the keys and values in the dictionary.
    #If the  configration option  is in the config it will fill excel cell. It will leave blank if the config doesnt exist.

    search = input("SNMP Comunity: ")

    success_fill = PatternFill(start_color='00FF00',
                                                end_color='00FF00',
                                                fill_type='solid')

    fail_fill = PatternFill(start_color='FF8080',
                                                end_color='FF8080',
                                                fill_type='solid')

    status = [ ]

    try:
        wb_file = "C:\Python\Book1.xlsx"
        workbook = load_workbook(wb_file)
        active_sheet = workbook.active
        active_sheet.protection = False

    except BaseException:
        print("Unknown file error")
    else:
        for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=10):
            for cell in row:
                if cell.value == 'Null':
                    print("\n")
                    print("Scan complete. Please review inventory sheet for send statuses. Red=Failed Connection, Green=Success", "No File =  User not Found")
                    print("\n")
                    time.sleep(2)
                    for item in status:
                        print(item)

                    try:
                        workbook.save("C:\Python\Book1.xlsx")
                        workbook.save(wb_file)
                        main()
                    except PermissionError:
                        print("Could not write to file. Please ensure the file isnt open and you have write permissions.")
                        main()
                else:
                    try:
                        m = manager.connect(host=cell.value, port=830, username="cisco", password="cisco", device_params={ 'name': 'csr'})

                        print("Device:", cell.value, "Session ID: ", m.session_id, " Connection: ", m.connected)
                        print("\n")

                        credential_config = open("C:\Python\XML_Filters\Filter_Config.xml").read()
                        config_data = m.get(credential_config)

                        config_details = xmltodict.parse(config_data.xml)["rpc-reply"]["data"]
                        parsed_config = config_details["native"]["snmp-server"]["community"]

                        for items in parsed_config:
                            snmp_comm = items
                            try:
                                if "name" in snmp_comm:
                                    print("Community: " + search + " exist")
                            except (KeyError, TypeError):
                                pass
                    except (KeyError, TypeError):
                        print("Community: " + search + " doesnt exist")
                        status.append(cell.value + " - Failed")
                        print("\n")
                        pass
                    except UnicodeError:
                        print("\n")
                        print("Invalid IP address. Please try again")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        pass
                    except ncclient.NCClientError:
                        print("\n")
                        print("Connection unsuccessful to " + cell.value)
                        cell.fill = fail_fill
                        status.append(cell.value + " - Failed")
                        pass
#####################################################

def search_tacacs():

    # Search config for a given configuration uses SEARCH functions to convert XML to dictionary.It then runs a for loop for the keys and values in the dictionary.
    #If the  configration option  is in the config it will fill excel cell. It will leave blank if the config doesnt exist.

    search = input("TACACS Group: ")

    success_fill = PatternFill(start_color='00FF00',
                                                end_color='00FF00',
                                                fill_type='solid')

    fail_fill = PatternFill(start_color='FF8080',
                                                end_color='FF8080',
                                                fill_type='solid')

    status= [ ]

    try:
        wb_file = "C:\Python\Book1.xlsx"
        workbook = load_workbook(wb_file)
        active_sheet = workbook.active
        active_sheet.protection = False

    except BaseException:
        print("Unknown file error")
    else:
        for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=10):
            for cell in row:
                if cell.value == 'Null':
                    print("\n")
                    print("Scan complete. Please review inventory sheet for send statuses. Red=Failed Connection, Green=Success", "No File =  User not Found")
                    print("\n")
                    time.sleep(2)
                    for item in status:
                        print(item)

                    try:
                        workbook.save("C:\Python\Book1.xlsx")
                        workbook.save(wb_file)
                        main()
                    except PermissionError:
                        print(
                            "Could not write to file. Please ensure the file isnt open and you have write permissions.")
                        main()
                else:
                    try:
                        m = manager.connect(host=cell.value, port=830, username="cisco", password="cisco", device_params={ 'name': 'csr'})


                        print("Device:", cell.value, "Session ID: ", m.session_id, " Connection: ", m.connected)
                        print("\n")

                        credential_config = open("C:\Python\XML_Filters\Filter_Config.xml").read()
                        config_data = m.get(credential_config)

                        config_details = xmltodict.parse(config_data.xml)["rpc-reply"]["data"]
                        parsed_config = config_details["native"]["tacacs"]["server"]

                        for items in parsed_config:
                            new_items = items
                            try:
                                for k, v in new_items.items():
                                    if v == search:
                                        cell.fill = success_fill
                                        status.append(cell.value + " - Success")
                                        print("\n")
                                        try:
                                            print("  Group: {} exist".format(new_items["name"]))
                                        except (KeyError, TypeError):
                                            pass
                                        try:
                                            print("  Server: {}".format(new_items["address"]["ipv4"]))
                                            print("\n")
                                        except (KeyError, TypeError):
                                            pass
                            except (KeyError, TypeError):
                                status.append(cell.value + " - Failed")
                                print("\n")
                                pass
                    except (KeyError, TypeError):
                        status.append(cell.value + " - Failed")
                        print("\n")
                        pass
                    except UnicodeError:
                        print("\n")
                        print("Invalid IP address. Please try again")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        pass
                    except ncclient.NCClientError:
                        print("\n")
                        print("Connection Unsuccessful")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        pass

#####################################################

def search_service_policy():

    # Search config for a given configuration uses SEARCH functions to convert XML to dictionary.It then runs a for loop for the keys and values in the dictionary.
    #If the  configration option  is in the config it will fill excel cell. It will leave blank if the config doesnt exist.

    search = input("Service Policy: ")

    success_fill = PatternFill(start_color='00FF00',
                                                end_color='00FF00',
                                                fill_type='solid')

    fail_fill = PatternFill(start_color='FF8080',
                                                end_color='FF8080',
                                                fill_type='solid')

    status = [ ]

    try:
        wb_file = "C:\Python\Book1.xlsx"
        workbook = load_workbook(wb_file)
        active_sheet = workbook.active
        active_sheet.protection = False

    except BaseException:
        print("Unknown file error")
    else:
        for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=10):
            for cell in row:
                if cell.value == 'Null':
                    print("\n")
                    print("Scan complete. Please review inventory sheet for send statuses. Red=Failed Connection, Green=Success", "No File =  User not Found")
                    time.sleep(2)
                    for item in status:
                        print(item)

                    try:
                        workbook.save("C:\Python\Book1.xlsx")
                        workbook.save(wb_file)
                        main()
                    except PermissionError:
                        print("Could not write to file. Please ensure the file isnt open and you have write permissions.")
                        main()
                else:
                    try:
                        m = manager.connect(host=cell.value, port=830, username="cisco", password="cisco", device_params={ 'name': 'csr'})

                        print("Device:", cell.value, "Session ID: ", m.session_id, " Connection: ", m.connected)
                        print("\n")

                        credential_config = open("C:\Python\XML_Filters\Filter_Config.xml").read()
                        config_data = m.get(credential_config)

                        config_details = xmltodict.parse(config_data.xml)["rpc-reply"]["data"]
                        parsed_config = config_details["native"]["interface"]["GigabitEthernet"]

                        for items in parsed_config:
                            new_items = items
                            try:
                                for k, v in new_items.items():
                                    if v == search:
                                        cell.fill = success_fill
                                        status.append(cell.value + " - Success")
                                        try:
                                            print("  GigabitEthernet: {}".format(new_items["name"]))
                                        except (KeyError, TypeError):
                                            print("Service-Policy " + search + " doesnt exist")
                                            status.append(cell.value + " - Failed")
                                            cell.fill = fail_fill
                                            print("\n")
                            except (KeyError, TypeError):
                                print("Service-Policy " + search + " doesnt exist")
                                status.append(cell.value + " - Failed")
                                cell.fill = fail_fill
                                print("\n")
                    except (KeyError, TypeError):
                        print("Service-Policy " + search + " doesnt exist")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        print("\n")
                        pass
                    except UnicodeError:
                        print("\n")
                        print("Invalid IP address. Please try again")
                        status.append(cell.value + " - Failed")
                        cell.fill = fail_fill
                        pass

#####################################################



def device_admin():

    print("Device Admin Menu")

    print("\n")
    print(" 1: Device Capabilities")
    print(" 2: View Config")
    print(" 3. Send Configuration")
    print(" 4: Main Menu")
    print("\n")

    config_selection = input("Please select an option:   ")

    if config_selection == "1":
        ncc_login("device", 830, "cisco", "cisco!", {'name': 'csr'})

        for c in m.server_capabilities:
            print(c)
        device_admin()

    elif config_selection == "2":

        config_selection = ' '
        while config_selection != '9':

            print("\n")
            print(" 1: OSPF")
            print(" 2: SNMPv2")
            print(" 3: Credentials")
            print(" 4: Interface(s)")
            print(" 5: QoS (multi device)")
            print(" 6: TACACS")
            print(" 7: Prefix-List")
            print(" 8: BGP")
            print(" 9: Main Menu")

            print("\n")
            config_selection = input("Please select an option:  ")

            if config_selection == "1":
                paramiko_login("show run | s ospf\n")
                break
            elif config_selection == "2":
                paramiko_login("show run | s snmp\n")
                break
            elif config_selection == "3":
                paramiko_login("show run | i user\n")
                break
            elif config_selection == "4":
                paramiko_login("show run | s interface\n ")
                break
            elif config_selection == "5":
                selection = input("Policy=map or Class-map: ")
                paramiko_login(" show run %s\n" % selection)
                break
            elif config_selection == "6":
                paramiko_login("show run | s tacacs\n")
                break
            elif config_selection == "7":
                paramiko_login("show ip prefix-list\n")
                break
            elif config_selection == "8":
                paramiko_login("show run | s bgp \n")
                break
            elif config_selection == "9":
                main()
                break
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

    elif config_selection == "3":
        select_configuration_send()
    elif config_selection == "4":
        main()
    else:
        device_admin()

def search_configurations():

    config_selection = ' '
    while config_selection != '5':

        print("\n")
        print("Configuration Search Menu")
        print("\n")

        print(" 1: Credentials")
        print(" 2: SNMP")
        print(" 3: TACACS")
        print(" 4: QoS")
        print(" 5: Main Menu")

        print("\n")
        config_selection = input("Please select an option:  ")

        if config_selection == "1":
            search_credentials()
        if config_selection == "2":
            search_snmp()
        if config_selection == "3":
            search_tacacs()
        if config_selection == "4":
            search_service_policy()
        elif config_selection == "5":
            main()
        else:
            print("\n")
            print("Invalid Selections")
            print("\n")

#############################################################################
def main():

    config_selection = ' '
    while config_selection != 'q':

        print("\n")
        print("Basic Network Programabiltiy and Automation Program")
        print("\n")

        print("\n")
        print(" 1: OSPF")
        print(" 2: SNMPv2 (multi device)")
        print(" 3: Credential (multi device)")
        print(" 4: Interface")
        print(" 5: DMVPN")
        print(" 6: QoS (multi device)")
        print(" 7. TACACS (multi device)")
        print(" 8. Prefix-List")
        print(" 9. BGP ")
        print(" 10: Device Admin")
        print(" 11: Search Configurations")
        print(" 12: FTP Inventory")
        print(" 13: Device Model/Serial")
        print("[q] (quit)")

        print("\n")

        config_selection = input("Please select an option:  ")

        if config_selection == "1":
            ospf_configuration()
        elif config_selection == "2":
            snmp_configuration()
        elif config_selection == "3":
            credentials_configuration()
        elif config_selection == "4":
            interface_configuration()
        elif config_selection == "5":
            dmvpn_configuration()
        elif config_selection == "6":
            qos_configuration()
        elif config_selection == "7":
            tacacs_configuration()
        elif config_selection == "8":
            prefix_configuration()
        elif config_selection == "9":
            bgp_configuration()
        elif config_selection == "10":
            device_admin()
        elif config_selection == "11":
            search_configurations()
        elif config_selection == "12":
            ftp_files()
        elif config_selection == "13":
            inventory()
        elif config_selection == "q":
            print("Exiting Program")
            print("\n")
        else:
            print("\n")
            print("Invalid Selection")
            print("\n")

    print("Thank you for using the Basic Network Programabiltiy and Automation Program")
    quit()

##############################################################################

def ospf_configuration():

    OSPF_file = "C:\Python\XML_Filters\Send_Config\OSPF_Send_Config.xml"
    Delete_Config = "C:\Python\XML_Filters\OSPF_Delete_ConfigTest.xml"

    print("OSPF Configuration:")

    config_selection = ' '
    while   config_selection != '5':

        try:

            print("\n")
            print("1. Add configuration" )
            print("2. Remove configuration")
            print("3. View OSPF Configuration")
            print("4. View OSPF Operational Status")
            print("5. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns:ios-ospf", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                router_element = xml.SubElement(native_element, "router")
                ospf_element = xml.SubElement(router_element, "ios-ospf:ospf")

                print("\n")
                print("OSPF Configuration")
                print("\n")
                print("Press enter to bypass configuration option")
                print("\n")

                process_input = input("Please Enter a OSPF Process: ")
                process_id = xml.SubElement(ospf_element, "ios-ospf:id")
                process_id.text = process_input

                rid_input = input("Please Enter a Router ID: ")
                rid_id = xml.SubElement(ospf_element, "ios-ospf:router-id")
                rid_id.text = rid_input

                print("\n")
                print("1. Add Network")
                print("2. Redistribution")
                print("3. Send Configuration")
                print("\n")

                config_selection = input("Please select an option:  ")

                if config_selection == "1":

                    network_leaf = xml.SubElement(ospf_element, "ios-ospf:network")

                    while True:
                        try:

                            network_input, mask_input, area_input = input("Please Enter a  OSPF network and mask: ").split()

                            ospf_network = network_input + "/" + mask_input
                            ipaddress.IPv4Network(ospf_network)

                        except ipaddress.AddressValueError:
                            print("\n")
                            print("Invalid Network Address")
                            print("\n")
                        except ipaddress.NetmaskValueError:
                            pass
                        else:

                            network_id = xml.SubElement(network_leaf, "ios-ospf:ip")
                            network_id.text = network_input

                            mask = xml.SubElement(network_leaf, "ios-ospf:mask")
                            mask.text = mask_input

                            area_id = xml.SubElement(network_leaf, "ios-ospf:area")
                            area_id.text = area_input

                            cleanup_empty_elements(root, OSPF_file)
                            view_config_send(OSPF_file)
                            break

                if config_selection == "2":

                    redis = xml.SubElement(ospf_element, "ios-ospf:redistribute")

                    while True:

                        print("\n")
                        print("1. Connected")
                        print("2. BGP")
                        print("3. Static")

                        print("\n")
                        config_selection = input("Please select an option:  ")
                        print("\n")

                        if config_selection == "1":

                            conn= xml.SubElement(redis, "ios-ospf:connected")
                            red_options =  xml.SubElement(conn, "ios-ospf:redist-options")
                            subnets = xml.SubElement(red_options, "ios-ospf:subnets")
                            subnets_2 = xml.SubElement(subnets, "ios-ospf:subnets")

                            tag_prefix = input("Please enter a tag: ")
                            tag = xml.SubElement(red_options, "ios-ospf:tag")
                            tag.text = tag_prefix

                            metric_type = input("Please enter a metric-type (default(2): ")
                            m_type = xml.SubElement(red_options, "ios-ospf:metric-type")
                            m_type.text = metric_type

                            red_map = input("Please enter a route-map: ")
                            route_map = xml.SubElement(red_options, "ios-ospf:route-map")
                            route_map.text = red_map
                            cleanup_empty_elements(root, OSPF_file)
                            view_config_send(OSPF_file)
                            break

                        if config_selection == "2":

                            while True:
                                bgp = xml.SubElement(redis, "ios-ospf:bgp")
                                as_num_input = int(input("Please enter a BGP AS: "))
                                AS_range = range(1, 65353)

                                if as_num_input not in AS_range:
                                    print(" Invalid AS")

                                else:

                                    AS_str = str(as_num_input)
                                    as_num =  xml.SubElement(bgp, "ios-ospf:as-number")
                                    as_num.text = AS_str

                                    red_options = xml.SubElement(bgp, "ios-ospf:redist-options")

                                    tag_prefix = input("Please enter a tag: ")
                                    tag = xml.SubElement(red_options, "ios-ospf:tag")
                                    tag.text = tag_prefix

                                    metric_type = input("Please enter a metric-type (default(2): ")
                                    m_type = xml.SubElement(red_options, "ios-ospf:metric-type")
                                    m_type.text = metric_type

                                    red_map = input("Please enter a route-map: ")
                                    route_map = xml.SubElement(red_options, "ios-ospf:route-map")
                                    route_map.text = red_map
                                    cleanup_empty_elements(root, OSPF_file)
                                    view_config_send(OSPF_file)
                                    break

                        if config_selection == "3":

                            static= xml.SubElement(redis, "ios-ospf:static")
                            red_options =  xml.SubElement(static, "ios-ospf:redist-options")
                            subnets = xml.SubElement(red_options, "ios-ospf:subnets")
                            subnets_2 = xml.SubElement(subnets, "ios-ospf:subnets")

                            tag_prefix = input("Please enter a tag: ")
                            tag = xml.SubElement(red_options, "ios-ospf:tag")
                            tag.text = tag_prefix

                            metric_type = input("Please enter a metric-type (default(2): ")
                            m_type = xml.SubElement(red_options, "ios-ospf:metric-type")
                            m_type.text = metric_type

                            red_map = input("Please enter a route-map: ")
                            route_map = xml.SubElement(red_options, "ios-ospf:route-map")
                            route_map.text = red_map
                            cleanup_empty_elements(root, OSPF_file)
                            view_config_send(OSPF_file)
                            break

                        else:
                            print("\n")
                            print("Invalid Selection")
                            print("\n")

                elif config_selection == "3":
                    cleanup_empty_elements(root, OSPF_file)
                    view_config_send(OSPF_file)
                    break

                else:
                    print("\n")
                    print("Invalid Selection")
                    print("\n")

            elif config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns:ios-ospf", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)

                config_selection_2 = input("Which proccess do you want to delete? ")

                router_element = xml.SubElement(native_element, "router")

                ospf_element = xml.SubElement(router_element, "ios-ospf:ospf")
                ospf_element.set("xc:operation", "delete")

                router_id = xml.SubElement(ospf_element, "ios-ospf:id")
                router_id.set("xc:operation", "delete")
                router_id.text = config_selection_2

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(Delete_Config, "wb") as fh:
                    tree.write(fh)

                send_single_configuration(Delete_Config)  # Call Function

            elif config_selection == "3":
                paramiko_login("show run | s ospf")
            elif config_selection == "4":

                print("\n")
                print("1. OSPF Neighbor")
                print("2. OSPF Status")

                print("\n")
                config_selection = input("Please select an option:  ")
                print("\n")

                while True:
                    if config_selection == "1":
                        paramiko_login("show ip ospf neighbor\n")
                    elif config_selection == "2":
                        proccess_id = input("OSPF Process: ")
                        paramiko_login("show ip ospf %s\n" % proccess_id)
                    else:
                        print("\n")
                        print("Invalid Selection")
                        print("\n")
            elif config_selection == "5":
                main()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. OSPF Configuration Menu")
            print("\n")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                ospf_configuration()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

    ###########################################################################

def snmp_configuration():

    SNMP_file = "C:\Python\XML_Filters\Send_Config\SNMPv2_Send_Config.xml"
    Delete_Config = "C:\Python\XML_Filters\SNMPv2_Delete_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add Configuration")
            print("2. Remove configuration")
            print("3. View SNMP Users")
            print("4. Main Menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                snmp_element = xml.Element("snmp-server")
                native_element.append(snmp_element)
                community_element = xml.Element("community")
                community_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-snmp")
                snmp_element.append(community_element)

                comm_input = input("Please Enter a Community: ")
                comm_id = xml.SubElement(community_element, "name")
                comm_id.text = comm_input

                question_1 = input("Do you want to enter an ACL? ")

                if question_1 == "yes":

                    acl_input = input("Please Enter An Access List: ")
                    acl_element = xml.SubElement(community_element, "access-list-name")
                    acl_element.text = acl_input

                    cleanup_empty_elements(root, SNMP_file)
                    view_config_send(SNMP_file)
                    break

                else:
                    cleanup_empty_elements(root, SNMP_file)
                    view_config_send(SNMP_file)
                    break

            elif config_selection == "2":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                snmp_element = xml.Element("snmp-server")
                native_element.append(snmp_element)
                community_element = xml.Element("community")
                community_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-snmp")
                community_element.set("operation", "delete")
                snmp_element.append(community_element)

                comm_input = input("Please Enter a Community to remove: ")
                comm_id = xml.SubElement(community_element, "name")
                comm_id.set("operation", "delete")
                comm_id.text = comm_input

                cleanup_empty_elements(root, Delete_Config)

                send_single_configuration(Delete_Config)

            elif config_selection == "3":
                paramiko_login("show run | s snmp\n")
            elif config_selection == "4":
                main()
                break
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. SNMP Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                snmp_configuration()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

###########################################################################

def credentials_configuration():

    Delete_User_Config = "C:\Python\XML_Filters\Credentials_Delete_Config.xml"
    Cred_Config = "C:\Python\XML_Filters\Send_Config\Credentials_Send_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add/Modify Configuration")
            print("2. Remove Configuration")
            print("3. View Users")
            print("4. Main Menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                username_element = xml.SubElement(native_element, "username")

                print("New User Configuration")

                user_input = input("Please Enter a Username: ")
                user_id = xml.SubElement(username_element, "name")
                user_id.text = user_input

                priv_input = input("Please Enter a Privilage Level: ")
                priv_element = xml.SubElement(username_element, "privilege")
                priv_element.text = priv_input

                while True:

                    print("\n")
                    print("1. Add Password")
                    print("2. Save Configuration")

                    config_selection = input("Please select an option:  ")

                    if config_selection == "1":

                        password_element = xml.SubElement(username_element, "password")

                        password_input = input("Please Enter Password: ")
                        pass_element = xml.SubElement(password_element, "password")
                        pass_element.text = password_input

                        cleanup_empty_elements(root, Cred_Config)
                        view_config_send(Cred_Config) # Call Function
                        break

                    elif config_selection == "2":

                        cleanup_empty_elements(root, Cred_Config)
                        view_config_send(Cred_Config)
                        break

                    else:
                        print("\n")
                        print("Invalid Selection")
                        print("\n")

            if config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                username_element = xml.SubElement(native_element, "username")
                username_element.set("xc:operation", "remove")

                question_1 = input("Which user do you want to delete? ")

                name_id = xml.SubElement(username_element, "name")
                name_id.set("xc:operation", "remove")
                name_id.text = question_1

                cleanup_empty_elements(root, Delete_User_Config)

                view_config_send(Delete_User_Config)

            if config_selection == "3":
                paramiko_login("show run | i user\n")
            if config_selection == "4":
                main()

            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Credential Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
                break
            if escape_1 == "2":
                credentials_configuration()
                break
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

    ################################################################################

def interface_configuration():

    int_file = "C:\Python\XML_Filters\Send_Config\Interface_Send_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add/Modify configuration")
            print("2. Remove Interface (Tunnel/Loopback")
            print("3. View Interface(s)")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")
            print("\n")

            if config_selection == "1":

                int_file = "C:\Python\XML_Filters\Send_Config\Interface_Send_Config.xml"
                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                int_element = xml.SubElement(native_element, "interface")

                readline.parse_and_bind("tab: complete")
                readline.set_completer(int_selection)

                while True:
                    print("Use tab to show options")
                    print("\n")
                    int_type = input("Please select an interface: ")

                    if int_type  ==  "GigabitEthernet"  or int_type ==  "FastEthernet" or int_type == "Loopback" or int_type == "Tunnel":

                        readline.parse_and_bind("tab: complete")
                        readline.set_completer(int_type_selection)

                        int_type_leaf = xml.SubElement(int_element, int_type)

                        interface_number = input("Please Enter An Interface number:  ")
                        int_name = xml.SubElement(int_type_leaf, "name")
                        int_name.text = interface_number

                        int_choice_1 = input("Please enter a description: ")
                        int_descrp = xml.SubElement(int_type_leaf, "description")
                        int_descrp.text = int_choice_1

                        while True:
                            try:

                                ip_input, mask_input = input("Please Enter A IP address and mask:  ").split()
                                ipaddress.IPv4Address(ip_input)

                            except ValueError:
                                print("\n")
                                print("Invalid Input")
                                print("\n")

                            else:

                                ip_leaf = xml.SubElement(int_type_leaf, "ip")
                                address_leaf = xml.SubElement(ip_leaf, "address")
                                primary_leaf = xml.SubElement(address_leaf, "primary")

                                address = xml.SubElement(primary_leaf, "address")
                                address.text = ip_input

                                mask = xml.SubElement(primary_leaf, "mask")
                                mask.text = mask_input

                                cleanup_empty_elements(root, int_file)
                                view_config_send(int_file)
                                cleanup_empty_elements(root, int_file)
                                view_config_send(int_file)
                                break

            if config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                int_element = xml.SubElement(native_element, "interface")

                readline.parse_and_bind("tab: complete")
                readline.set_completer(int_selection)

                while True:
                    print("Use tab to show options")
                    print("\n")
                    int_type = input("Enter an interface type: ")

                    if int_type == "GigabitEthernet" or int_type == "FastEthernet" or int_type == "Loopback" or int_type == "Tunnel":

                        readline.parse_and_bind("tab: complete")
                        readline.set_completer(int_type_selection)

                        int_type = xml.SubElement(int_element, int_type)
                        int_type.set("xc:operation", "remove")

                        interface_number = input("Please Enter An Interface number:  ")
                        int_name = xml.SubElement(int_type, "name")
                        int_name.set("xc:operation", "remove")
                        int_name.text = interface_number

                        cleanup_empty_elements(root, int_file)
                        view_config_send(int_file)
                        break

            if config_selection == "3":
                paramiko_login("show run | s interface\n ")
            elif config_selection == "4":
                main()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Interface Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
                break
            if escape_1 == "2":
                interface_configuration()
                break
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

################################################################################

def dmvpn_configuration():

    dmvpn_file = "C:\Python\XML_Filters\Send_Config\DMVPN_Send_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add/Modify Configuration")
            print("2. View Interface(s)")
            print("3. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                int_element = xml.SubElement(native_element, "interface")

                int_type_leaf = xml.SubElement(int_element, "Tunnel")

                interface_name = input("Please Enter An Interface: Tunnel: ")
                int_name = xml.SubElement(int_type_leaf, "name")
                int_name.text = interface_name

                ip_leaf = xml.SubElement(int_type_leaf, "ip")

                int_choice_1 = input("Please enter a description: ")
                int_descrp = xml.SubElement(int_type_leaf, "description")
                int_descrp.text = int_choice_1

                print("\n")
                print("1. Add New DMVPN Tunnel")
                print("2. Add/Modiy NHRP")
                print("3. Add/Modify NHS")
                print("4. Add/Modify NHS Mapping")
                print("5. Change Tunnel Source")
                print("6. Save Configuration")

                config_selection = input("Please select an option: ")

                if config_selection == "1":

                    while True:
                        try:
                            ip_input, mask_input = input("Please Ente a IP address and mask: ").split()
                            ipaddress.IPv4Address(ip_input)

                        except ipaddress.AddressValueError:
                            print("\n")
                            print("Invalid Network Address")
                            print("\n")
                        except ipaddress.NetmaskValueError:
                            print("\n")
                            print("Invalid Wilcard")
                            print("\n")
                        except ValueError:
                            print("\n")
                            print("Invalid Input")
                            print("\n")
                        else:

                            ip_leaf = xml.SubElement(int_type_leaf, "ip")
                            address_leaf = xml.SubElement(ip_leaf, "address")
                            primary_leaf = xml.SubElement(address_leaf, "primary")

                            address = xml.SubElement(primary_leaf, "address")
                            address.text = ip_input

                            mask = xml.SubElement(primary_leaf, "mask")
                            mask.text = mask_input

                            readline.parse_and_bind("tab: complete")
                            readline.set_completer(ospf_net_type_selection)

                            ospf_leaf = xml.SubElement(ip_leaf, "ospf")
                            ospf_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")

                            ospf_net_type = input("Please Enter OSPF network Type: ")
                            ospf_type = xml.SubElement(ospf_leaf, "network")
                            ospf_type.text = ospf_net_type

                            cleanup_empty_elements(root, dmvpn_file)
                            view_config_send(dmvpn_file)
                            break

                elif config_selection == "2":

                    nhrp_leaf = xml.SubElement(ip_leaf, "nhrp")
                    nhrp_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-nhrp")

                    nhrp_input1 = input("Please Enter NHRP auth ")
                    nhrp_auth = xml.SubElement(nhrp_leaf, "authentication")
                    nhrp_auth.text = nhrp_input1

                    nhrp_input2 = input("Please Enter NHRP group: ")
                    nhrp_group = xml.SubElement(nhrp_leaf, "group")
                    nhrp_group.text = nhrp_input2

                    nhrp_input3 = input("Please Enter NHRP holdtime: ")
                    nhrp_hold = xml.SubElement(nhrp_leaf, "holdtime")
                    nhrp_hold.text = nhrp_input3

                    nhrp_input8 = input("Please Enter NHRP network ID: ")
                    nhrp_ID = xml.SubElement(nhrp_leaf, "network-id")
                    nhrp_ID.text = nhrp_input8

                    nhrp_shortcut = xml.SubElement(nhrp_leaf, "shortcut")
                    nhrp_redirect = xml.SubElement(nhrp_leaf, "redirect")

                    cleanup_empty_elements(root, dmvpn_file)
                    view_config_send(dmvpn_file)
                    break

                elif config_selection == "3":
                    while True:
                        try:

                            nhs_ip= input("Please Enter NHRP NHS: ")
                            ipaddress.IPv4Address(nhs_ip)

                        except ipaddress.AddressValueError:
                            print("\n")
                            print("Invalid Network Address")
                            print("\n")
                        except ValueError:
                            print("\n")
                            print("Invalid Input")
                            print("\n")
                        else:

                            nhrp_leaf = xml.SubElement(ip_leaf, "nhrp")
                            nhrp_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-nhrp")
                            nhs_container = xml.SubElement(nhrp_leaf, "nhs")
                            nhs_leaf = xml.SubElement(nhs_container, "ipv4")
                            nhrp_nhs = xml.SubElement(nhs_leaf, "ipv4")
                            nhrp_nhs.text = nhs_ip

                            config_selection_2 = input("Do you want to add a NHS prioirty (yes/no?) ")

                            while True:
                                if config_selection_2 == "yes":

                                    nhrp_pri_el = xml.SubElement(nhs_leaf, "priority")
                                    nhrp_range = xml.SubElement(nhrp_pri_el, "pri-range")

                                    nhrp_prio= input("Please Enter NHS Priority: ")
                                    nhrp_range= xml.SubElement(nhrp_range, "pri-range")
                                    nhrp_range.text = nhrp_prio

                                    cleanup_empty_elements(root, dmvpn_file)
                                    view_config_send(dmvpn_file)
                                    break

                                if config_selection_2 =="no":

                                    cleanup_empty_elements(root, dmvpn_file)
                                    view_config_send(dmvpn_file)
                                    break

                                else:
                                    print("\n")
                                    print("Invalid Input")
                                    print("\n")

                elif config_selection == "4":
                    while True:
                        try:

                            hub_tun = input("Please Enter Hub Tunnel IP:  ")
                            ipaddress.IPv4Address(hub_tun)

                            hub_nbma = input("Which NBMA IP did the tunnel change for:  ")
                            ipaddress.IPv4Address(hub_nbma)

                        except ipaddress.AddressValueError:
                            print("\n")
                            print("Invalid IP Address")
                            print("\n")
                        except ValueError:
                            print("\n")
                            print("Invalid Input")
                            print("\n")
                        else:

                            nhrp_leaf = xml.SubElement(ip_leaf, "nhrp")
                            nhrp_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-nhrp")
                            nhrp_map_leaf = xml.SubElement(nhrp_leaf, "map")

                            nhrp_dest_leaf = xml.SubElement(nhrp_map_leaf, "dest-ipv4")

                            nhrp_hub = xml.SubElement(nhrp_dest_leaf, "dest-ipv4")
                            nhrp_hub.text = hub_tun

                            nhrp_nbma1 = xml.SubElement(nhrp_dest_leaf, "nbma-ipv4")

                            nhrp_hub = xml.SubElement(nhrp_nbma1, "nbma-ipv4")
                            nhrp_hub.text = hub_nbma

                            multicast_container = xml.SubElement(nhrp_map_leaf, "multicast")

                            nhrp_hub = xml.SubElement(multicast_container, "nbma_ipv4")
                            nhrp_hub.text = hub_nbma

                            cleanup_empty_elements(root, dmvpn_file)
                            view_config_send(dmvpn_file)
                            break

                elif config_selection == "5":

                    readline.parse_and_bind("tab: complete")
                    readline.set_completer(int_selection)

                    tun_leaf = xml.SubElement(int_type_leaf, "tunnel")
                    tun_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-tunnel")

                    source_input, int_type = input("Please Enter a tunnel source: ").split()
                    tun_source = xml.SubElement(tun_leaf, "source")
                    tun_source.text = source_input + " " +  int_type

                    cleanup_empty_elements(root, dmvpn_file)
                    view_config_send(dmvpn_file)
                    break

                else:

                    cleanup_empty_elements(root, dmvpn_file)
                    view_config_send(dmvpn_file)
                    break

            if config_selection == "2":
                tun_num = input("Tunnel Number: ")
                paramiko_login("show run interface tunnel %s\n" % tun_num)
            if config_selection == "3":
                main()
                break
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. DMVPN Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                dmvpn_configuration()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

###################################################################3

def qos_configuration():

    classmap_file = "C:\Python\XML_Filters\Send_Config\QoS_Send_Config.xml"
    policy_map_file = "C:\Python\XML_Filters\Send_Config\PolicyMap_Send_Config.xml"
    serv_policy_file = "C:\Python\XML_Filters\Send_Config\PolicyMap_Shape_Send_Config.xml"
    int_policy_map_file = "C:\Python\XML_Filters\Send_Config\PolicyMap_Interface_Shape_Send_Config.xml"

    config_selection = ' '
    while config_selection != '6':

        try:
            print("\n")
            print("1. Configure Class-map")
            print("2. Configure Policy-map")
            print("3. Confgure Policy-map (shaping)")
            print("4. Apply QoS (interface)")
            print("5. View QoS")
            print("6. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":


                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                policy_element = xml.Element("policy")
                native_element.append(policy_element)
                class_element = xml.Element("class-map")
                class_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
                policy_element.append(class_element)

                print("\n")
                print("Create class-map")
                print("\n")

                global class1_input
                class1_input = input("Please Enter class-map name: ")
                class_id = xml.SubElement(class_element, "name")
                class_id.text = class1_input

                readline.parse_and_bind("tab: complete")
                readline.set_completer(qos_match_selection)

                class1_1_input = input("Please Enter class-map type (match-any/all): ")
                prematch_element = xml.SubElement(class_element, "prematch")
                prematch_element.text = class1_1_input

                match_element = xml.SubElement(class_element, "match")

                while class1_1_input  not in {'match-all', 'match-any'}:
                    class1_1_input = input("Please Enter class-map type (match-any/all): ")

                else:
                    try:
                        while True:
                            print("Please add match statements")
                            print("\n")
                            print("Press CTRL+C to escape at any time")
                            print("\n")

                            tag_input, tag_value= str(input("Please Enter tag type: ")),  int(input("Please enter a tag value: "))

                            COS_range = range(1, 7)
                            DSCP_range = (0,10, 12, 14, 18, 20, 22, 26, 28, 30, 24, 26, 38, 8, 16, 34, 24, 32, 40, 40, 56, 46)
                            AF_range = (21, 22, 23, 31, 32, 33, 41, 42, 43)

                            if tag_input == "cs" and tag_value in COS_range or tag_input == "dscp" and tag_value in DSCP_range or tag_input == "af" \
                                    and tag_value in AF_range:

                                tag_value_str = str(tag_value)
                                match_1_element = xml.SubElement(match_element, "tag")
                                match_1_element.text = tag_input + tag_value_str

                                cleanup_empty_elements(root, classmap_file)

                            else:
                                print("\n")
                                print("Invalid Input")
                                print("\n")
                    except ValueError:
                        print("\n")
                        print("Invalid Input")
                        print("\n")


            if config_selection == "2":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                policy_element = xml.Element("policy")
                native_element.append(policy_element)
                policy_map_element = xml.Element("policy-map")
                policy_map_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
                policy_element.append(policy_map_element)

                pol_map_input = input("Please create policy map name: ")
                pol_map_name = xml.SubElement(policy_map_element, "name")
                pol_map_name.text = pol_map_input

                print("\n")
                print("Policy-map Configuration")
                print("\n")

                while True:

                    print("\n")
                    print("Press CTRL+C to escape at any time")
                    print("\n")

                    pol_class1_element = xml.SubElement(policy_map_element, "class")

                    class_to_queue = input("Please create a  queue: ")
                    pol_name_1 = xml.SubElement(pol_class1_element, "name")
                    pol_name_1.text = class_to_queue

                    action_list_element = xml.SubElement(pol_class1_element, "action-list")

                    readline.parse_and_bind("tab: complete")
                    readline.set_completer(qos_action_selection)

                    band_prio = input("Please Enter class-map type (bandwidth/priority): ")
                    action_type_element = xml.SubElement(action_list_element, "action-type")
                    action_type_element.text =band_prio

                    priority_container = xml.SubElement(action_list_element, band_prio)

                    bandwidth_input = input("Please enter bandwith percent: ")
                    percent_element = xml.SubElement(priority_container, "percent")
                    percent_element.text = bandwidth_input

                    cleanup_empty_elements(root, policy_map_file)

            if config_selection == "3":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                policy_element = xml.Element("policy")
                native_element.append(policy_element)
                policy_map_element = xml.Element("policy-map")
                policy_map_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
                policy_element.append(policy_map_element)

                pol_map_5_input = input("Please create policy map name: ")
                pol_map_5_name = xml.SubElement(policy_map_element, "name")
                pol_map_5_name.text = pol_map_5_input

                pol_class5_element = xml.SubElement(policy_map_element, "class")

                nest_input = input("Please enter a nested class (class-default): ")
                nested_name = xml.SubElement(pol_class5_element, "name")
                nested_name.text = nest_input

                action_list_5_element = xml.SubElement(pol_class5_element, "action-list")

                while True:

                    readline.parse_and_bind("tab: complete")
                    readline.set_completer(qos_shape_pol_selection)

                    shape_or_police = input("Shape or Police: : ")
                    action_type_5_element = xml.SubElement(action_list_5_element, "action-type")
                    action_type_5_element.text = shape_or_police

                    shape_element = xml.SubElement(action_list_5_element, shape_or_police)
                    average_element = xml.SubElement(shape_element, "average")

                    bandwidth_5_input = input("Please enter bandwidth in bits: ")
                    percent_5_element = xml.SubElement(average_element, "bit-rate")
                    percent_5_element.text = bandwidth_5_input

                    action_list_element = xml.SubElement(pol_class5_element, "action-list")
                    action_type_element = xml.SubElement(action_list_element, "action-type")
                    action_type_element.text = "service-policy"

                    child_pol_input = input("Please enter a child policy map: ")
                    action_list_5 = xml.SubElement(action_list_element, "service-policy")
                    action_list_5.text = child_pol_input

                    cleanup_empty_elements(root, serv_policy_file)
                    view_config_send(serv_policy_file)

            if config_selection == "4":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)

                int_element = xml.SubElement(native_element, "interface")

                int_type = input("Please enter interface type: ")
                int_type_element= xml.SubElement(int_element, int_type)

                int_num = input("Please enter interface number:")
                int_num_element = xml.SubElement(int_type_element, "name")
                int_num_element.text = int_num

                service_1_element = xml.SubElement(int_type_element, "service-policy")
                service_1_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")

                output_1_policy = input("Please enter service policy: ")
                service_1_output = xml.SubElement(service_1_element, "output")
                service_1_output.text = output_1_policy

                cleanup_empty_elements(root, int_policy_map_file)

                view_config_send(int_policy_map_file)

            if config_selection == "5":
                selection = input("Polciy=map or Class-map: ")
                paramiko_login(" show run %s\n" % selection)
            if config_selection == "6":
                main()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Send Class-map Configuration")
            print("3. Send Policy-map Configuration")
            print("4. QoS Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                view_config_send(classmap_file)
            if escape_1 == "3":
                view_config_send(policy_map_file)
            if escape_1 == "4":
                qos_configuration()
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

def tacacs_configuration():

    tacacs_file = "C:\Python\XML_Filters\Send_Config\TACACS_Send_Config.xml"

    selection = " "
    while selection != "4":

        try:

            print("\n")
            print("1. Add/Modify Configuration")
            print("2. Remove Configuration")
            print("3. View TACACS")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                tacacs_elem = xml.SubElement(native_element, "tacacs")
                tac_server = xml.SubElement(tacacs_elem, "server")
                tac_server.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-aaa")

                print("\n")
                print("Configure TACACS Server:")
                print("\n")

                tac_name_input = input("Please enter a TACACS group: ")
                tac_name = xml.SubElement(tac_server, "name")
                tac_name.text = tac_name_input

                print("\n")
                print("1. Add/Change TACACS Server")
                print("2. Main Menu")

                config_selection = input("Please select an option: ")

                if config_selection == "1":
                    while True:
                        try:
                            tac_ipv4_input = input("Please enter a TACACS server IP: ")
                            ipaddress.ip_address(tac_ipv4_input)

                        except ValueError:

                            print("\n")
                            print("Invalid IP")
                            print("\n")

                        else:

                            tac_address = xml.SubElement(tac_server, "address")
                            tac_ipv4 = xml.SubElement(tac_address, "ipv4")
                            tac_ipv4.text = tac_ipv4_input

                            tac_key_elem = xml.SubElement(tac_server, "key")

                            tac_key_input = input("Please enter a TACACS key: ")
                            tac_key = xml.SubElement(tac_key_elem, "key")
                            tac_key.text = tac_key_input

                            cleanup_empty_elements(root, tacacs_file)
                            view_config_send(tacacs_file)
                            break

                if config_selection == "2":
                    main()
                    break

                else:
                    print("\n")
                    print("Invalid Selection")
                    print("\n")

            elif config_selection == "2":

                while True:

                    print("\n")
                    print("Configure TACACS Server:")
                    print("\n")

                    print("\n")
                    print("1. Remove TACACS Group")
                    print("2. Main Menu")
                    print("\n")
                    print("Press CTRL+C to escape at any time")
                    print("\n")

                    config_selection = input("Please select an option:  ")

                    if config_selection == "1":

                        del_tacacs = "C:\Python\XML_Filters\Send_Config\TACACS_Delete_Config.xml"
                        root = xml.Element("config")
                        root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                        root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                        native_element = xml.Element("native")
                        native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                        root.append(native_element)

                        tacacs_elem = xml.SubElement(native_element, "tacacs")
                        tac_server = xml.SubElement(tacacs_elem, "server")
                        tac_server.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-aaa")
                        tac_server.set("xc:operation", "remove")

                        tac_name_input = input("Please enter a TACACS group: ")
                        tac_name = xml.SubElement(tac_server, "name")
                        tac_name.set("xc:operation", "remove")
                        tac_name.text = tac_name_input

                        cleanup_empty_elements(root, del_tacacs)
                        view_config_send(del_tacacs)
                        break

                    elif config_selection == "2":
                        main()
                        break
                    else:
                        print("\n")
                        print("Invalid Selection")
                        print("\n")

            elif config_selection == "3":
                paramiko_login("show run | s tacacs\n")
            elif config_selection == "4":
                main()
                break
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")


        except KeyboardInterrupt:

                    print("\n")
                    print("1. Main Menu")
                    print("2. TACACS Configuration Menu")

                    escape_1 = input("What menu do you want to escape to?")

                    if escape_1 == "1":
                        main()
                    if escape_1 == "2":
                        tacacs_configuration()
                    else:
                        print("\n")
                        print("Invalid Selection")
                        print("\n")

def prefix_configuration():

    prefix_file = "C:\Python\XML_Filters\Send_Config\Prefix_Send_Config.xml"
    prefix_del = "C:\Python\XML_Filters\Prefix_Delete_Config.xml"


    while True:
        try:
            while True:

                print("\n")
                print("1. Add configuration")
                print("2. Remove/Modify configuration")
                print("3. View Prefix-list")
                print("4. Main menu")

                config_selection = input("Please select an option:  ")

                if config_selection == "1":

                    root = xml.Element("config")
                    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    native_element = xml.Element("native")
                    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                    root.append(native_element)

                    print("\n")
                    print("Configure Prefix-List:")
                    print("\n")

                    prefix = xml.SubElement(native_element, "ip")
                    prefix_list = xml.SubElement(prefix, "prefix-list")
                    prefixes = xml.SubElement(prefix_list, "prefixes")

                    prefix_name_input = input("Please enter a prefix-list name: ")
                    prefix_name = xml.SubElement(prefixes, "name")
                    prefix_name.text = prefix_name_input

                    while config_selection == "1":
                        while True:
                            try:
                                print("\n")
                                print("Press CTRL+C to escape at any time")
                                print("\n")

                                # Auto increase prefix list sequence by 5 starting 5, up to 1000. Currently this block of code will not through an expcetion if a prefix length is not entered
                                # The configuration will through an expcetion when sent to the device.

                                for i in range(5, 1000, 5):

                                    seq_input = i

                                    seq = xml.SubElement(prefixes, "seq")
                                    num = xml.SubElement(seq, "no")
                                    num.text = str(seq_input)

                                    readline.parse_and_bind("tab: complete")
                                    readline.set_completer(permit_deny_selection)

                                    per_deny_input = input("Permit ir Deny?")
                                    permit_deny = xml.SubElement(seq, per_deny_input)

                                    prefix_input = input("Please enter a prefix: ")
                                    ipaddress.IPv4Network(prefix_input)

                                    ip = xml.SubElement(permit_deny, "ip")
                                    ip.text = prefix_input

                                    cleanup_empty_elements(root, prefix_file)

                            except ValueError:
                                print("\n")
                                print("Invalid Input")
                                print("\n")
                            else:
                                cleanup_empty_elements(root, prefix_file)


                elif config_selection == "2":

                    root = xml.Element("config")
                    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    native_element = xml.Element("native")
                    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                    root.append(native_element)

                    print("\n")
                    print("Modify/Delete  Prefix-List:")
                    print("\n")

                    prefix = xml.SubElement(native_element, "ip")
                    prefix_list = xml.SubElement(prefix, "prefix-list")
                    prefixes = xml.SubElement(prefix_list, "prefixes")
                    prefixes.set("xc:operation", "remove")

                    while True:

                        print("\n")
                        print("1. Remove Prefix List")
                        print("2. Remove/Modify Prefix Sequence")

                        config_selection = input("Please select an option:  ")

                        if config_selection == "1":

                            prefix_name_input = input("Please enter a prefix-list name: ")
                            prefix_name = xml.SubElement(prefixes, "name")
                            prefix_name.text = prefix_name_input
                            prefix_name.set("xc:operation", "remove")

                            cleanup_empty_elements(root, prefix_del)
                            view_config_send(prefix_del)
                            break

                        if config_selection == "2":

                            prefix = xml.SubElement(native_element, "ip")
                            prefix_list = xml.SubElement(prefix, "prefix-list")
                            prefixes = xml.SubElement(prefix_list, "prefixes")

                            prefix_name_input = input("Please enter a prefix-list name: ")
                            prefix_name = xml.SubElement(prefixes, "name")
                            prefix_name.text = prefix_name_input

                            seq = xml.SubElement(prefixes, "seq")
                            seq.set("xc:operation", "remove")

                            seq_input = input("Please enter a seq number: ")
                            num = xml.SubElement(seq, "no")
                            num.set("xc:operation", "remove")
                            num.text = seq_input

                            cleanup_empty_elements(root, prefix_del)
                            view_config_send(prefix_del)
                            break

                        else:
                            print("\n")
                            print("Invalid Selection")
                            print("\n")


                elif config_selection == "3":
                    paramiko_login("show ip prefix-list\n")
                elif config_selection == "4":
                    main()
                else:
                    print("\n")
                    print("Invalid Selection")
                    print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Prefix-List Configuration Menu")
            print("3. Send Prefix- List Configuration ")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                prefix_configuration()
            if escape_1 == "3":
                view_config_send(prefix_file)
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

def bgp_configuration():


    bgp_file = "C:\Python\XML_Filters\Send_Config\BGP_Send_Config.xml"
    bgp_del = "C:\Python\XML_Filters\BGP_Delete_Config.xml"

    root = xml.Element("config")
    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
    native_element = xml.Element("native")
    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
    root.append(native_element)
    router_elem = xml.SubElement(native_element, "router")
    bgp_elem = xml.SubElement(router_elem, "bgp")
    bgp_elem.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-bgp")

    while True:

        AS_input = int(input("Please enter a local AS: "))
        AS_range = range(1, 65353)

        if AS_input not in AS_range:
            print("invalid AS")

        else:

            AS_str = str(AS_input) # CONVERTS INTIGER TO STRING TO WIRTE TO XML TREE. FUNCTION WONT WRITE INTIGERS TO ELEEMENTS
            AS_id = xml.SubElement(bgp_elem, "id")
            AS_id.text = AS_str
            cleanup_empty_elements(root, bgp_file)

            while True:

                print("\n")
                print("1. Add Neighbor")
                print("2. Add Network Statement")
                print("3. Redistribtion")
                print("4. View Configuration")
                print("5. View BGP Operational Status")
                print("\n")


                config_selection = input("Please select an option:  ")

                if config_selection =="1":

                    bgp_neigh = xml.SubElement(bgp_elem, "neighbor")

                    neigh_id= input("Please enter a neighbor: ")
                    id = xml.SubElement(bgp_neigh, "id")
                    id.text = neigh_id

                    while True:

                        neigh_AS = int(input("Please enter a neighbor AS: "))
                        AS_range = range(1, 65353)

                        if neigh_AS not in AS_range:
                            print("invalid AS")

                        else:

                            AS_str = str(neigh_AS) # CONVERTS INTIGER TO STRING TO WIRTE TO XML TREE. FUNCTION WONT WRITE INTIGERS TO ELEEMENTS
                            AS_id = xml.SubElement(bgp_neigh, "remote-as")
                            AS_id.text = AS_str
                            cleanup_empty_elements(root, bgp_file)

                            soft_reconf = xml.SubElement(bgp_neigh, "soft-reconfiguration")
                            soft_reconf.text = "inbound"

                            cleanup_empty_elements(root, bgp_file)
                            view_config_send(bgp_file)


                elif config_selection == "2":
                    bgp_network = xml.SubElement(bgp_elem, "network")

                    while True:
                        try:
                            bgp_net_input = input("Please enter a network: ")
                            ipaddress.IPv4Network(bgp_net_input)

                            bgp_mask_input= input("Please enter network mask: ")
                            ipaddress.IPv4Network(bgp_mask_input)

                        except ValueError:
                            print("\n")
                            print("Invalid Configuration")
                            print("\n")
                        except ipaddress.IPv4Network:
                            print("\n")
                            print("Invalid Network")
                            print("\n")
                        else:

                            bgp_num = xml.SubElement(bgp_network, "number")
                            bgp_num.text = bgp_net_input

                            bgp_mask = xml.SubElement(bgp_network, "mask")
                            bgp_mask.text = bgp_mask_input

                            cleanup_empty_elements(root, bgp_file)
                            view_config_send(bgp_file)

                elif config_selection == "3":

                    redis = xml.SubElement(bgp_elem, "redistribute")

                    while True:

                        print("\n")
                        print("1. Connected")
                        print("2. OSPF")

                        print("\n")
                        config_selection = input("Please select an option:  ")
                        print("\n")

                        if config_selection == "1":

                            redis = xml.SubElement(redis, "connected")

                            red_map = input("Please enter a route-map: ")
                            route_map= xml.SubElement(redis, "route-map")
                            route_map.text = red_map
                            cleanup_empty_elements(root, bgp_file)
                            view_config_send(bgp_file)
                            break

                        if config_selection == "2":

                            redis = xml.SubElement(redis, "ospf")

                            red_map = input("Please enter a route-map: ")
                            route_map = xml.SubElement(redis, "route-map")
                            route_map.text = red_map
                            cleanup_empty_elements(root, bgp_file)
                            view_config_send(bgp_file)
                            break

                        else:
                            print("\n")
                            print("Invalid Selection")
                            print("\n")
                elif config_selection == "4":
                    paramiko_login("show run | s bgp\n")

                elif config_selection == "5":

                    print("\n")
                    print("1. Neighbor")
                    print("2. BGP Table")

                    print("\n")
                    config_selection = input("Please select an option:  ")
                    print("\n")

                    while True:
                        if config_selection == "1":
                            paramiko_login("show ip bgp summary\n")
                        elif config_selection == "2":
                            paramiko_login("show ip bgp\n")
                        else:
                            print("\n")
                            print("Invalid Selection")
                            print("\n")
            else:
                print("\n")
                print("Invalid Selection")
                print("\n")

def inventory():

    #Get serial nuber and model of a device. IT works with with ISR 4331 but not 3850. After reviewing 3850 config i dont see a model listed with xml file.

    filename = "C:\Python\XML_Filters\Serial_Model.xml"

    root = xml.Element("filter")
    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
    native_element = xml.Element("native")
    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
    root.append(native_element)
    license_elem = xml.SubElement(native_element, "license")

    tree = xml.ElementTree(root)
    with open(filename, "wb") as fh:
        tree.write(fh)

    print("\n")
    print("1. Multi-device")
    print("2. Single device")

    while True:
        print("\n")
        config_selection = input("Please select an option:  ")
        print("\n")

        if config_selection == "1":

            success_fill = PatternFill(start_color='00FF00',
                                       end_color='00FF00',
                                       fill_type='solid')

            fail_fill = PatternFill(start_color='FF8080',
                                    end_color='FF8080',
                                    fill_type='solid')

            try:
                wb_file = "C:\Python\Book1.xlsx"
                workbook = load_workbook(wb_file)
                active_sheet = workbook.active
                active_sheet.protection = False

            except BaseException:
                print("Unknown file error")
            else:
                for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=10):
                    for cell in row:
                        if cell.value == 'Null':
                            print("\n")
                            print(
                                "Scan complete. Please review inventory sheet for send statuses. Red=Failed Connection, Green=Success",
                                "No File =  User not Found")
                            time.sleep(2)

                            try:
                                workbook.save("C:\Python\Book1.xlsx")
                                workbook.save(wb_file)
                                main()
                            except PermissionError:
                                print(
                                    "Could not write to file. Please ensure the file isnt open and you have write permissions.")
                                main()
                        else:
                            try:
                                global m
                                m = manager.connect(host=cell.value, port=830, username="cisco", password="cisco",
                                                    device_params={'name': 'csr'})

                                print("\n")
                                print("Device:", cell.value, "Session ID: ", m.session_id, " Connection: ", m.connected)
                                print("\n")

                                serial_path = open("C:\Python\XML_Filters\Serial_Model.xml").read()
                                serial_get = m.get(serial_path)
                                lic_details = xmltodict.parse(serial_get.xml)["rpc-reply"]["data"]
                                lic_config = lic_details["native"]["license"]

                                try:
                                    print("  Model: {}".format(lic_config["udi"]["pid"]))
                                except (KeyError, TypeError):
                                    pass
                                try:
                                    print("  Serial: {}".format(lic_config["udi"]["sn"]))
                                except (KeyError, TypeError):
                                    print("Unable to get model/serial")
                                    pass
                            except (KeyError, TypeError):
                                print("Unable to get model/serial" )
                                print("\n")
                                pass
                            except UnicodeError:
                                print("\n")
                                print("Invalid IP address. Please try again")
                                cell.fill = fail_fill
                                pass
                            except ncclient.NCClientError:
                                print("\n")
                                print("Connection Unsuccessful")
                                cell.fill = fail_fill
                                pass

        elif config_selection == "2":

            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

            serial_path = open("C:\Python\XML_Filters\Serial_Model.xml").read()
            serial_get = m.get(serial_path)
            lic_details = xmltodict.parse(serial_get.xml)["rpc-reply"]["data"]
            lic_config = lic_details["native"]["license"]
            print("")
            print("Device Details:")

            try:
                print("  Model: {}".format(lic_config["udi"]["pid"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Serial: {}".format(lic_config["udi"]["sn"]))
                break
            except (KeyError, TypeError):
                print("Can't find serial")
                pass
        else:
            print("\n")
            print("Invalid Selection")
            print("\n")

if __name__ == '__main__':

    main()

