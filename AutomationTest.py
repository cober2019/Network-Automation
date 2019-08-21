from ncclient import manager
import  ncclient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xml.etree.ElementTree as xml
import xml.dom.minidom as dom
import lxml.etree as ET
import xmltodict
import pathlib
from socket import gaierror
import ftplib
import time
from ncclient.operations import RPCError

def ftp_files():

    # DOWNLOADS FILE FROM AN FTP SERVER AND SAVES THEM TO A LOCAL DIRECTORY

    filename = "Book1.xlsx"
    dest_file = open("C:\Python" + "\\" + filename, 'wb')
    dest_path =  "C:\Python" + "\\" + filename

    print(" 1: FTP Download")
    print(" 2: Main Menu")
    print("\n")
    print("Press CTRL+C to escape at any time")
    print("\n")

    print("\n")
    question_1 = input("Please select an option: ")
    print("\n")

    if question_1 == "1":
        print("\n")
        ftp_server = input("Please enter an FTP server: ")
        print("\n")

        ftp = ftplib.FTP(ftp_server)
        ftp.login("admin", "1234")
        ftp.cwd("/")

        try:
            ftp.retrbinary("RETR " + filename, dest_file.write, 1024)
            print("File Downloaded: Path:  " + dest_path)
            time.sleep(2)
            ftp.quit()
            main()

        except ftplib.error_perm:
            print("Please check the following - Filepath, credentials, Filename")
            ftp_files()
        except TimeoutError:
            print("Could not connect to FTP server. Please check connectivity")
            ftp_files()
        except UnicodeError:
            print("Invalid IP address. Please re-enter")
            ftp_files()

    if question_1 == "2":
        main()

    else:
        print("Invalid input")
        ftp_files()

#################################################################

def select_configuration_send():

    # CONFIGURATION FILES CAN BE VIEWED FROM THIS DIRECTORY AND SENT. THIS CAN BE USED TO SEND FILES THAT INTIALLY FAILED

    print(" 1: File Select")
    print(" 2: Main Menu")

    print("\n")
    question_1 = input("Please select an option: ")
    print("\n")

    if question_1 == "1":

        dir_files = []
        for p in pathlib.Path("C:\Python\XML_Filters\Send_Config").iterdir():
            if p.is_file():
                print(p)
                dir_files.append(p)

        file = input("Please enter a filename: ")

        try:
            config_file = open("C:\Python\XML_Filters\Send_Config" + "\\" +  file).read()
            print(dom.parseString(str(config_file)).toprettyxml())

            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
            send_payload = m.edit_config(config_file, target="running")
            save_configuration()
            main()

        except ncclient.operations.rpc.RPCError:
            print("\n")
            print("Invalid XML Element value. Please review your configurations")
            time.sleep(2)
            print("\n")

        except FileNotFoundError:
            print("\n")
            print("File not found!")
            time.sleep(2)
            print("\n")
            select_configuration_send()

    if question_1 == "2":
        main()

    else:
        print("Invalid input, please try again")

#######################################################################################


def save_configuration():

    # SAVES RUNNING CONFIG TO STARTUP CONFIG AFTER CONFIGURATION SEND

    save_payload = """
    		<cisco-ia:save-config xmlns:cisco-ia="http://cisco.com/yang/cisco-ia"/>
            """
    try:
        response = m.dispatch(ET.fromstring(save_payload))
        print("Configuration Saved")
        main()
    except ValueError:
        print("Error")
        main()

########################################################################################

def view_config_send(file):

    # DEFINES SINGLE DEVICE CONIFGURATION OR MULTI DEVICE SEND. YOU ARE ALSO ABLE TO CANCLE A SEND AS WELL

    print("\n")
    print ("Configuration: ")
    print("\n")

    Interface_config = open(file=file).read()
    print(dom.parseString(str(Interface_config)).toprettyxml())

    while True:

        print(" 1: Send Single Device Configuration")
        print(" 2: Send Multi  Device Configuration")
        print(" 3: Cancel Configuration Send")

        print("\n")
        question_1 = input("Please select an option: ")
        print("\n")

        if question_1 == "1":
            send_single_configuration(file=file)

        if question_1 == "2":
            send__multi_configuration(file=file)

        if question_1 == "3":
            print("\n")
            print("Configuration Canceled")
            print("\n")
            main()
        else:
            print("\n")
            print("Invalid input, please try again")
            print("\n")



######################################################################


def ncc_login(host, port, username, password, device_params): # Log into device via NCC client

    #NCCLIENT LOGIN
    try:
        global m
        global device
        print("\n")
        device = input("Please Enter a device IP address: ")
        print("\n")
        print("\n")
        print("Press CTRL+C to escape at any time")
        print("\n")

        try:
            m = manager.connect(host=device, port=port, username=username, password=password, device_params=device_params)

            print("\n")
            print("Device:", device, "Session ID: ", m.session_id, " Connection: ",m.connected)
            print("\n")
        except (UnicodeError):
            print("Invalid IP address. Insure IP format is correct")
            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
        except ncclient.NCClientError:
            print("Connection Timeout. Please check connectivity and NETCONF configuration.")
            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
        except gaierror:
            print("Invalid IP address. Insure IP format is correct")
            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})
    except KeyboardInterrupt:
        main()



################################################################################

def send__multi_configuration(file):

    #READS DEVICE FROM AND EXCEL SPREAD AND FILLS THE CELL WITH EITHER GREEN OR RED. A GREEN FILLED CELL IS A SUCCESS, RED IS A FAILURE. THESE ARE FILLED BASED ON AN EXCEPTION DURING THE JOB

    success_fill = PatternFill(start_color='00FF00',
                                                end_color='00FF00',
                                                fill_type='solid')

    fail_fill = PatternFill(start_color='FF8080',
                                                end_color='FF8080',
                                                fill_type='solid')

    wb_file = "C:\Python\Book1.xlsx"
    workbook = load_workbook(wb_file)
    active_sheet = workbook.active
    active_sheet.protection = False

    for row in active_sheet.iter_rows(min_row=1, max_col=1, max_row=5):
        for cell in row:
            print(cell.value)

            if cell.value == 'Null':
                print("\n")
                print("Job complete. Please review inventory sheet for send statuses. Red=Failed, Green=Success")
                time.sleep(2)

                try:
                    workbook.save("C:\Python\Book1.xlsx")
                    workbook.save(wb_file)
                    main()
                except PermissionError:
                    print("Could not write to file. Please ensure the file isnt open and you have write permissions.")
                    main()
            else:

                try:
                    x = manager.connect(host=cell.value, port=830, username="cisco", password="cisco", device_params={
                        'name': 'csr'})

                    print("\n")
                    print("Device:", cell.value, "Session ID: ", x.session_id, " Connection: ", x.connected)
                    print("\n")

                    config_file = open(file=file).read()
                    send_payload = x.edit_config(config_file, target="running")

                    save_payload = """
                                                    <cisco-ia:save-config xmlns:cisco-ia="http://cisco.com/yang/cisco-ia"/>
                                              """
                    try:
                        response = x.dispatch(ET.fromstring(save_payload))
                    except ValueError:
                        print("\n")
                        print("Configuration Saved")
                        cell.fill = success_fill
                        continue
                    except AttributeError:
                        print("\n")
                        print("Connection Unsucessful")
                        cell.fill = fail_fill
                        pass
                    except RPCError:
                        print("\n")
                        print("Payload Error")
                        cell.fill = fail_fill
                        pass
                except (UnicodeError):
                    print("\n")
                    print("Invalid IP address. Please try again")
                    cell.fill = fail_fill
                    pass
                except ncclient.NCClientError:
                    print("\n")
                    print("Please review configuration.")
                    cell.fill = fail_fill
                    pass

    ##################################################################################

def send_single_configuration(file):

        # USED FOR SINGLE DEVICE CONFIGURATION. ONCE CONFIGURATION IS COMPLETE, YOU WILL BE RETURNED TO THE MENU FROM WHERE YOU CONFIGURING. THIS IS BASED OF THE FILENAME OF THE CONFIGURATION

        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        try:
            config_file = open(file=file).read()
            send_payload = m.edit_config(config_file, target="running")
            save_configuration()

            print("\n")
            print("Configuration Complete!")
            print("\n")
            main()

            if "Credentials" in file:
                view_users()

            if "SNMP" in file:
                view_snmp_users()

            if "OSPF" in file:
                view_ospf()

            if "Interface" in file:
                view_interfaces()

            if "QoS" in file:
                view_qos()

            if "TACACS" in file:
                view_tacacs()

            if "Prefix" in file:
                view_prefix_list()

        except ValueError:
            print("Configuration Saved")
            main()
        except AttributeError:
            print("Connection Unsucessful")
            pass
        except RPCError:
            pass
        except (UnicodeError):
            print("Invalid IP address. Please try again")
            pass
        except ncclient.NCClientError:
            print("Please review configuration")
            pass



######################################################################################################

def view_qos():

    # VIEW CLASS-MAPS, POLICY-MAPS, SERVICE POLICIES. XML VIEW IN ONLY AVAIBLE FOR NOW.

    classmap_file = "C:\Python\XML_Filters\QoS_Get_Config.xml"
    root = xml.Element("filter")
    native_element = xml.Element("native")
    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
    root.append(native_element)
    policy_element = xml.Element("policy")
    native_element.append(policy_element)
    class_element = xml.Element("class-map")
    class_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
    policy_element.append(class_element)

    tree = xml.ElementTree(root)
    with open(classmap_file, "wb") as fh:
        tree.write(fh)

    print("\n")
    print("Interface Selection Menu")

    print("\n")
    print(" 1: View Class-maps")
    print(" 2: View Policy-maps")
    print(" 3: Interface Policy")
    print(" 4: QoS Configuration")
    print(" 5: Main Menu")
    print("\n")

    config_selection = input("Please select an option:  ")

    if config_selection == "1":

        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        qos_config = open("C:\Python\XML_Filters\QoS_Get_Config.xml").read()
        qos_get = m.get(qos_config)
        qos_details = xmltodict.parse(qos_get.xml)["rpc-reply"]["data"]

        qos_config = qos_details["native"]["policy"]["class-map"]
        print("")
        print("Qos Details:")

        try:
            print("  Class-map: {}".format(qos_config["name"]))
        except (KeyError, TypeError):
            pass
        try:
            print("  match-:{}".format(qos_config["prematch"]))
        except (KeyError, TypeError):
            pass
        try:
            print("  Tag: {}".format(qos_config["match"]["dscp"]))
        except (KeyError, TypeError):
            pass

        for item in qos_config:
            qos_filter = item

            try:
                print("")
                if "name" in qos_filter:
                    try:
                        print("  Class-map: {}".format(qos_filter["name"]))
                    except (KeyError, TypeError):
                        pass
                if "prematch" in qos_filter:
                    try:
                        print("  Match Type: {}".format(qos_filter["prematch"]))
                    except (KeyError, TypeError):
                        pass
                if "match" in qos_filter:
                    try:
                        print("  Tag: {}".format(qos_filter["match"]["dscp"]))
                    except (KeyError, TypeError):
                        pass
            except (KeyError, TypeError):
                pass
                print("No Class-map configured")

        print("\n")
        qos_configuration()
        print("\n")

    if config_selection == "2":

        policy_map_file = "C:\Python\XML_Filters\PolicyMap_Get_Config.xml"

        root = xml.Element("filter")
        native_element = xml.Element("native")
        native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
        root.append(native_element)
        policy_element = xml.Element("policy")
        native_element.append(policy_element)
        policy_map_element = xml.Element("policy-map")
        policy_map_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
        policy_element.append(policy_map_element)

        tree = xml.ElementTree(root)
        with open(policy_map_file, "wb") as fh:
            tree.write(fh)

        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        pol_config = open("C:\Python\XML_Filters\PolicyMap_Get_Config.xml").read()
        pol_get = m.get(pol_config)
        print(dom.parseString(str(pol_get)).toprettyxml())

        print("\n")
        qos_configuration()
        print("\n")

    if config_selection == "3":
        view_interfaces()
    if config_selection == "4":
        qos_configuration()
    elif config_selection == "5":
        main()

########################################################################

def view_interfaces():

        filename = "C:\Python\XML_Filters\Interface_Get_Config.xml"
        root = xml.Element("filter")
        root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
        root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
        native_element = xml.Element("native")
        native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
        root.append(native_element)
        int_element = xml.SubElement(native_element, "interface")

        tree = xml.ElementTree(root)
        with open(filename, "wb") as fh:
            tree.write(fh)

        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        intf_config = open("C:\Python\XML_Filters\Interface_Get_Config.xml").read()
        intf_get = m.get(intf_config)
        intf_details = xmltodict.parse(intf_get.xml)["rpc-reply"]["data"]

        print("\n")
        print("Interface Selection Menu")

        print("\n")
        print(" 1: Tunnel")
        print(" 2: Loopback")
        print(" 3: GigabitEthernet")
        print(" 4: Interface Configuration")
        print(" 5: QoS Configuration")
        print(" 6: Main Menu")
        print("\n")

        config_selection = input("Please select an option:  ")

        if config_selection == "1":

            tun_config = intf_details["native"]["interface"]["Tunnel"]
            print("")
            print("Interface Details:")

            try:
                print("  Tunnel: {}".format(tun_config["name"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Description: {}".format(tun_config["description"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  VRF: {}".format(tun_config["vrf"]["forwarding"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  IP: {}".format(tun_config["ip"]["address"]["primary"]["address"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Mask: {}".format(tun_config["ip"]["address"]["primary"]["mask"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  NHRP Auth: {}".format(tun_config["ip"]["nhrp"]["authentication"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  NHRP Group: {}".format(tun_config["ip"]["nhrp"]["group"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  NHRP Holtime: {}".format(tun_config["ip"]["nhrp"]["holdtime"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  NHS: {}".format(tun_config["ip"]["nhrp"]["map"]["dest-ipv4"]["dest-ipv4"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  NHS NMBA: {}".format(tun_config["ip"]["nhrp"]["map"]["dest-ipv4"]["nbma-ipv4"]["nbma-ipv4"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  OSPF Net Type: {}".format(tun_config["ip"]["nhrp"]["ospf"]["network"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  NHRP Network ID: {}".format(tun_config["ip"]["nhrp"]["map"]["dest-ipv4"]["network-id"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Tunnel Source: {}".format(tun_config["tunnel"]["source"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Service Policy: {}".format(tun_config["service-policy"]["output"]))
            except (KeyError, TypeError):
                pass

            for item in tun_config:
                tun_filter = item

                try:
                    print("")
                    if "name" in tun_filter:
                        try:
                            print("  Tunnel: {}".format(tun_filter["name"]))
                        except (KeyError, TypeError):
                            pass
                    if "description" in tun_filter:
                        try:
                            print("  Description: {}".format(tun_filter["description"]))
                        except (KeyError, TypeError):
                            pass
                    if "vrf" in tun_filter:
                        try:
                            print("  VRF: {}".format(tun_filter["vrf"]["forwarding"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  IP: {}".format(tun_filter["ip"]["address"]["primary"]["address"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  Mask: {}".format(tun_filter["ip"]["address"]["primary"]["mask"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  NHRP Auth: {}".format(tun_filter["ip"]["nhrp"]["authentication"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  NHRP Group: {}".format(tun_filter["ip"]["nhrp"]["group"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  NHRP Holtime: {}".format(tun_filter["ip"]["nhrp"]["holdtime"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  NHS: {}".format(tun_filter["ip"]["nhrp"]["map"]["dest-ipv4"]["dest-ipv4"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  NHS NMBA: {}".format(tun_filter["ip"]["nhrp"]["map"]["dest-ipv4"]["nbma-ipv4"]["nbma-ipv4"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  NHRP Network ID: {}".format(tun_filter["ip"]["nhrp"]["map"]["dest-ipv4"]["network-id"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  OSPF Net Type: {}".format(tun_filter["ip"]["nhrp"]["ospf"]["network"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  Tunnel Source: {}".format(tun_filter["tunnel"]["source"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in tun_filter:
                        try:
                            print("  Service-Policy {}".format(tun_filter["service-policy"]["output"]))
                        except (KeyError, TypeError):
                            pass
                except (KeyError, TypeError):
                    pass
                    print("No Tunnel Interfaces")

            print("\n")
            interface_configuration()
            print("\n")

        if config_selection == "2":

            loop_config = intf_details["native"]["interface"]["Loopback"]
            print("")
            print("Interface Details:")

            try:
                print("  Loopback: {}".format(loop_config["name"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Description: {}".format(loop_config["description"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  VRF: {}".format(loop_config["vrf"]["forwarding"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  IP: {}".format(loop_config["ip"]["address"]["primary"]["address"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Mask: {}".format(loop_config["ip"]["address"]["primary"]["mask"]))
            except (KeyError, TypeError):
                pass

            for item in loop_config:
                loop_filter = item

                try:
                    print("")
                    if "name" in loop_filter:
                        try:
                            print("  Loopback: {}".format(loop_filter["name"]))
                        except (KeyError, TypeError):
                            pass
                    if "description" in loop_filter:
                        try:
                            print("  Description: {}".format(loop_filter["description"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in loop_filter:
                        try:
                            print("  IP: {}".format(loop_filter["ip"]["address"]["primary"]["address"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in loop_filter:
                        try:
                            print("  Mask: {}".format(loop_filter["ip"]["address"]["primary"]["mask"]))
                        except (KeyError, TypeError):
                            pass
                except :
                    print("No Loopback Interfaces")
                    pass

            print("\n")
            interface_configuration()
            print("\n")

        if config_selection == "3":


            gig_config = intf_details["native"]["interface"]["GigabitEthernet"]
            print("")
            print("Interface Details:")
            try:
                print("  GigabitEthernet: {}".format(gig_config["name"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Description: {}".format(gig_config["description"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  IP: {}".format(gig_config["ip"]["address"]["primary"]["address"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Mask: {}".format(gig_config["ip"]["address"]["primary"]["mask"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Service-Policy  {}".format(gig_config["service-policy"]["output"]))
            except (KeyError, TypeError):
                pass

            for item in gig_config :
                int_filter = item

                try:
                    print("")
                    if "name" in int_filter:
                        try:
                            print("  GigabitEthernet: {}".format(int_filter["name"]))
                        except (KeyError, TypeError):
                            pass
                    if "vrf" in int_filter:
                        try:
                            print("  VRF: {}".format(int_filter["vrf"]["forwarding"]))
                        except (KeyError, TypeError):
                            pass
                    if "description" in int_filter:
                        try:
                            print("  Description: {}".format(int_filter["description"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in int_filter:
                        try:
                            print("  IP: {}".format(int_filter["ip"]["address"]["primary"]["address"]))
                        except (KeyError, TypeError):
                            pass
                    if "ip" in int_filter:
                        try:
                            print("  Mask: {}".format(int_filter["ip"]["address"]["primary"]["mask"]))
                        except (KeyError, TypeError):
                            pass
                    if "service-policy" in int_filter:
                        try:
                            print("  Service-Policy {}".format(int_filter["service-policy"]["output"]))
                        except (KeyError, TypeError):
                            pass
                except:
                    print("No Gigabit Interfaces")
                    pass

            print("\n")
            interface_configuration()
            print("\n")

        if config_selection == "4":
            interface_configuration()
        if config_selection == "5":
            qos_configuration()
        if config_selection == "6":
            main()
        else:
            print("\n")
            print("Invalid input, please try again")


###################################################################################

def view_ospf():

        filename = "C:\Python\XML_Filters\OSPF_Get_Config.xml"
        root = xml.Element("filter")
        native_element = xml.Element("native")
        native_element.set("xmlns:ios-ospf", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")
        native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
        root.append(native_element)
        router_element = xml.SubElement(native_element, "router")
        ospf_element = xml.SubElement(router_element, "ios-ospf:ospf")

        tree = xml.ElementTree(root)
        with open(filename, "wb") as fh:
            tree.write(fh)

        print("\n")
        print("OSPF Menu")
        print("\n")

        print(" 1: View OSPF Configurations")
        print(" 2: Configure OSPF")
        print(" 3: Main Menu")

        print("\n")
        config_selection = input("Please select an option:  ")
        print("\n")

        if config_selection == "1":

            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

            ospf_config = open("C:\Python\XML_Filters\OSPF_Get_Config.xml").read()
            ospf_get = m.get(ospf_config)
            ospf_details = xmltodict.parse(ospf_get.xml)["rpc-reply"]["data"]
            ospf_config = ospf_details["native"]["router"]["ospf"]

            try:
                print("  Process: {}".format(ospf_config["id"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Router ID: {}".format(ospf_config["router-id"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Network: {}".format(ospf_config["network"]["ip"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Wildcard: {}".format(ospf_config["network"]["mask"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Area: {}".format(ospf_config["network"]["area"]))
            except (KeyError, TypeError):
                pass

            for item in ospf_config:
                ospf_filter = item

                ospf_config = open("C:\Python\XML_Filters\OSPF_Get_Config.xml").read()
                ospf_get = m.get(ospf_config)
                ospf_details = xmltodict.parse(ospf_get.xml)["rpc-reply"]["data"]
                ospf_config = ospf_details["native"]["router"]["ospf"]

                print("")
                print("OSPF Details:")
                if "id" in ospf_filter:
                    try:
                        print("  Process: {}".format(ospf_filter["id"]))
                    except (KeyError, TypeError):
                        pass
                if "router-id" in ospf_filter:
                    try:
                        print("  Router ID: {}".format(ospf_filter["router-id"]))
                    except (KeyError, TypeError):
                        pass
                if "network" in ospf_filter:
                    try:
                        print("  Network: {}".format(ospf_filter["network"]["ip"]))
                    except (KeyError, TypeError):
                        pass
                if "network" in ospf_filter:
                    try:
                        print("  Wildcard: {}".format(ospf_filter["network"]["mask"]))
                    except (KeyError, TypeError):
                        pass
                if "network" in ospf_filter:
                    try:
                        print("  Area: {}".format(ospf_filter["network"]["area"]))
                    except (KeyError, TypeError):
                        pass

            print("\n")
            view_ospf()
            print("\n")

        if config_selection == "2":
            ospf_configuration()
        if config_selection == "3":
            main()
        else:
                print("\n")
                print("Invalid input, please try again")


###################################################################################

def view_tacacs():

    tacacs_file = "C:\Python\XML_Filters\TACACS_Get_Config.xml"
    root = xml.Element("filter")
    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
    native_element = xml.Element("native")
    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
    root.append(native_element)

    tacacs_elem = xml.SubElement(native_element, "tacacs")


    tree = xml.ElementTree(root)
    with open(tacacs_file, "wb") as fh:
        tree.write(fh)

    print("TACACS  Menu")

    print("\n")
    print(" 1: View TACACS Groups")
    print(" 2: Configure TCACS")
    print(" 3: Main Menu")

    print("\n")

    config_selection = str(input("Please select an option:  "))

    if config_selection == "1":
        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        TACACS_config = open("C:\Python\XML_Filters\TACACS_Get_Config.xml").read()
        TACACS_get = m.get(TACACS_config)
        TACACS_details = xmltodict.parse(TACACS_get.xml)["rpc-reply"]["data"]
        TACACS_config = TACACS_details["native"]["tacacs"]["server"]

        try:
            print("  Group: {}".format(TACACS_config["name"]))
        except (KeyError, TypeError):
            pass
        try:
            print("  Server: {}".format(TACACS_config["address"]["ipv4"]))
        except (KeyError, TypeError):
            pass
        try:
            print("  Key: {}".format(TACACS_config["key"]["key"]))
        except (KeyError, TypeError):
            pass

        for item in TACACS_config:
            tacacs_conf = item

            print("")
            print("TACACS Details:")
            try:
                if "name" in tacacs_conf:
                    print("  Group: {}".format(tacacs_conf["name"]))
            except (KeyError, TypeError):
                pass
            if "address" in tacacs_conf:
                try:
                    print("  Server: {}".format(tacacs_conf["address"]["ipv4"]))
                except (KeyError, TypeError):
                    pass
            if "key" in tacacs_conf:
                try:
                    print("  Key: {}".format(tacacs_conf["key"]["key"]))
                except (KeyError, TypeError):
                    pass

        print("\n")
        view_tacacs()
        print("\n")

        if config_selection == "2":
            tacacs_configuration()
        if config_selection == "3":
            main()
        else:
            print("\n")
            print("Invalid input, please try again")

def view_prefix_list():

    filename = "C:\Python\XML_Filters\Prefix_Get_Config.xml"
    root = xml.Element("filter")
    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
    native_element = xml.Element("native")
    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
    root.append(native_element)

    prefix = xml.SubElement(native_element, "ip")
    prefix_list = xml.SubElement(prefix, "prefix-list")

    tree = xml.ElementTree(root)
    with open(filename, "wb") as fh:
        tree.write(fh)

    print("Prefix-List  Menu")

    print("\n")
    print(" 1: View Prefix-List")
    print(" 2: Configure Prefix-List")
    print(" 3: Main Menu")

    print("\n")

    config_selection = str(input("Please select an option:  "))

    if config_selection == "1":
        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        prefix_config = open("C:\Python\XML_Filters\Prefix_Get_Config.xml").read()
        prefix_get = m.get(prefix_config)

        print(dom.parseString(str(prefix_get)).toprettyxml())

        print("\n")
        view_prefix_list()
        print("\n")

    if config_selection == "2":
        prefix_configuration()
    if config_selection == "3":
        main()
    else:
        print("\n")
        print("Invalid input, please try again")



def view_snmp_users():

        filename = "C:\Python\XML_Filters\SNMPv2_Get_Config.xml"
        root = xml.Element("filter")
        native_element = xml.Element("native")
        native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
        root.append(native_element)
        snmp_element = xml.Element("snmp-server")
        native_element.append(snmp_element)
        community_element = xml.Element("community")
        community_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-snmp")
        snmp_element.append(community_element)

        tree = xml.ElementTree(root)
        with open(filename, "wb") as fh:
            tree.write(fh)

        print("SNMP  Menu")

        print("\n")
        print(" 1: View SNMP Users")
        print(" 2: Configure SNMP Users")
        print(" 3: Main Menu")

        print("\n")

        config_selection = str(input("Please select an option:  "))

        if config_selection== "1":

            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

            SNMPv2_config = open("C:\Python\XML_Filters\SNMPv2_Get_Config.xml").read()
            SNMP_get = m.get(SNMPv2_config)
            SNMP_details = xmltodict.parse(SNMP_get.xml)["rpc-reply"]["data"]
            SNMP_config = SNMP_details["native"]["snmp-server"]["community"]

            try:
                print("  Community: {}".format(SNMP_config["name"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Access: {}".format(SNMP_config["access-list-name"]))
            except (KeyError, TypeError):
                pass

            for item in SNMP_config:
                snmp_comm = item

                print("")
                print("SNMP Details:")
                try:
                    if "name" in snmp_comm:
                        print("  Community: {}".format(snmp_comm["name"]))
                except (KeyError, TypeError):
                    pass
                try:
                    if "access-list-name" in snmp_comm:
                        print("  Access: {}".format(snmp_comm["access-list-name"]))
                except (KeyError, TypeError):
                    pass

            print("\n")
            view_snmp_users()
            print("\n")

        if config_selection == "2":
            snmp_configuration()
        if config_selection == "3":
            main()
        else:
            print("\n")
            print("Invalid input, please try again")

###################################################################################


def view_users():

        filename = "C:\Python\XML_Filters\Credentials_Get_Config.xml"
        root = xml.Element("filter")
        root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
        native_element = xml.Element("native")
        native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
        root.append(native_element)
        username_element = xml.SubElement(native_element, "username")

        tree = xml.ElementTree(root)
        with open(filename, "wb") as fh:
            tree.write(fh)

        print("User Menu")

        print("\n")
        print(" 1: View Users")
        print(" 2: Configure Users")
        print(" 3: Main Menu")
        print("\n")

        config_selection = input("Please select an option:  ")

        if config_selection == "1":

            ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

            credential_config = open("C:\Python\XML_Filters\Credentials_Get_Config.xml").read()
            config_data = m.get(credential_config)

            cred_details = xmltodict.parse(config_data.xml)["rpc-reply"]["data"]
            cred_config = cred_details["native"]["username"]

            print("")
            print("Username Details:")
            try:
                    print("  Username: {}".format(cred_config["name"]))
            except (KeyError, TypeError):
                pass
            try:
                    print("  Priv: {}".format(cred_config["privilege"]))
            except (KeyError, TypeError):
                pass
            try:
                    print("  Encryption: {}".format(cred_config["password"]["encryption"]))
            except (KeyError, TypeError):
                pass
            try:
                    print("  Password: {}".format(cred_config["password"]["password"]))
            except (KeyError, TypeError):
                pass
            try:
                    print("  Enable Encryption: {}".format(cred_config["secret"]["encryption"]))
            except (KeyError, TypeError):
                pass
            try:
                print("  Enable Password: {}".format(cred_config["secret"]["secret"]))
                print("\n")
            except (KeyError, TypeError):
                pass

            for item in cred_config:
                user = item

                print("")
                print("Username Details:")
                try:
                    if "name" in user:
                        print("  Username: {}".format(user["name"]))
                except (KeyError, TypeError):
                    pass
                try:
                    if "privilege" in user:
                        print("  Priv: {}".format(user["privilege"]))
                except (KeyError, TypeError):
                    pass
                try:
                    if "password" in user:
                        print("  Encryption: {}".format(user["password"]["encryption"]))
                except (KeyError, TypeError):
                    pass
                try:
                    if "password" in user:
                        print("  Password: {}".format(user["password"]["password"]))
                except (KeyError, TypeError):
                    pass
                try:
                    if "secret" in user:
                        print("  Enable Encryption: {}".format(user["secret"]["encryption"]))
                except (KeyError, TypeError):
                    pass
                try:
                    if "secret" in user:
                        print("  Enable Password: {}".format(user["secret"]["secret"]))
                        print("\n")
                except (KeyError, TypeError):
                    pass

            print("\n")
            view_users()
            print("\n")

        if config_selection == "2":
            credentials_configuration()
        if config_selection == "3":
            main()
        else:
            print("\n")
            print("Invalid selection")
            print("\n")

###########################################################################

def device_admin():

    print("Device Admin Menu")

    print("\n")
    print(" 1: Device Capabilities")
    print(" 2: View Config")
    print(" 3. Send Configuration")
    print(" 4: Main Menu")
    print("\n")

    config_selection = input("Please select an option?  ")

    if config_selection == "1":
        ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

        for c in m.server_capabilities:
            print(c)
        device_admin()

    elif config_selection == "2":

        config_selection = ' '
        while config_selection != '8':

            print("\n")
            print(" 1: OSPF")
            print(" 2: SNMPv2")
            print(" 3: Credentials")
            print(" 4: Interface(s)")
            print(" 5: QoS (multi device)")
            print(" 6: TACACS")
            print(" 7: Prefix-List")
            print(" 8: Main Menu")

            print("\n")
            config_selection = input("Please select an option:  ")

            if config_selection == "1":
                view_ospf()
            elif config_selection == "2":
                view_snmp_users()
            elif config_selection == "3":
                view_users()
            elif config_selection == "4":
                view_interfaces()
            elif config_selection == "5":
                view_qos()
            elif config_selection == "6":
                view_tacacs()
            elif config_selection == "7":
                view_prefix_list()
            elif config_selection == "8":
                main()
            else:
                print("\n")
                print("Invalid selection")
                print("\n")

    elif config_selection == "3":
        select_configuration_send()
    elif config_selection == "4":
        main()
    else:
        device_admin()


#############################################################################
def main():

    config_selection = ' '
    while config_selection != 'q':

        print("\n")
        print("Basic Network Programabiltiy and Automation Program")
        print("\n")

        print("\n")
        print(" 1: OSPF")
        print(" 2: SNMPv2 (multi device)")
        print(" 3: Credential (multi device)")
        print(" 4: Interface")
        print(" 5: DMVPN")
        print(" 6: QoS (multi device)")
        print(" 7. TACACS (multi device)")
        print(" 8. Prefix-List")
        print(" 9: Device Admin")
        print(" 10: FTP Inventory")
        print("[q] (quit)")

        print("\n")

        config_selection = input("Please select an option:  ")

        if config_selection == "1":
            ospf_configuration()
        elif config_selection == "2":
            snmp_configuration()
        elif config_selection == "3":
            credentials_configuration()
        elif config_selection == "4":
            interface_configuration()
        elif config_selection == "5":
            dmvpn_configuration()
        elif config_selection == "6":
            qos_configuration()
        elif config_selection == "7":
            tacacs_configuration()
        elif config_selection == "8":
            prefix_configuration()
        elif config_selection == "9":
            device_admin()
        elif config_selection == "10":
            ftp_files()
        elif config_selection == "q":
            print("Exiting Program")
            print("\n")
        else:
            print("Invalid Selection")

    print("Thank you for using the Basic Network Programabiltiy and Automation Program")
    quit()

##############################################################################

def ospf_configuration():

    OSPF_file = "C:\Python\XML_Filters\Send_Config\OSPF_Send_ConfigTest.xml"
    Delete_Config = "C:\Python\XML_Filters\OSPF_Delete_ConfigTest.xml"

    print("OSPF Configuration:")

    config_selection = ' '
    while   config_selection != '4':

        try:

            print("\n")
            print("1. Add configuration" )
            print("2. Remove configuration")
            print("3. View OSPF Configuration")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns:ios-ospf", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                router_element = xml.SubElement(native_element, "router")
                ospf_element = xml.SubElement(router_element, "ios-ospf:ospf")

                Delete_Config = "C:\Python\XML_Filters\OSPF_Delete_Config.xml"

                print("\n")
                print("New Routing Configuration")
                print("\n")

                process_input = input("Please Enter a OSPF Process: ")
                process_id = xml.SubElement(ospf_element, "ios-ospf:id")
                process_id.text = process_input

                rid_input = input("Please Enter a Router ID: ")
                rid_id = xml.SubElement(ospf_element, "ios-ospf:router-id")
                rid_id.text = rid_input

                network_leaf = xml.SubElement(ospf_element, "ios-ospf:network")

                network_input = input("Please Enter a Network ID: ")
                network_id = xml.SubElement(network_leaf, "ios-ospf:ip")
                network_id.text = network_input

                mask_input = input("Please Enter a Wildcard: ")
                mask = xml.SubElement(network_leaf, "ios-ospf:mask")
                mask.text = mask_input

                area_input = input("Please Enter a Area: ")
                area_id = xml.SubElement(network_leaf, "ios-ospf:area")
                area_id.text = area_input

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(OSPF_file, "wb") as fh:
                    tree.write(fh)

                view_config_send(OSPF_file)

            elif config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns:ios-ospf", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)

                config_selection_2 = input("Which proccess do you want to delete? ")

                router_element = xml.SubElement(native_element, "router")

                ospf_element = xml.SubElement(router_element, "ios-ospf:ospf")
                ospf_element.set("xc:operation", "delete")

                router_id = xml.SubElement(ospf_element, "ios-ospf:id")
                router_id.set("xc:operation", "delete")
                router_id.text = config_selection_2

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(Delete_Config, "wb") as fh:
                    tree.write(fh)

                send_single_configuration(Delete_Config)  # Call Function

            elif config_selection == "3":
                view_ospf()
            elif config_selection == "4":
                main()
            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. OSPF Configuration Menu")
            print("\n")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                ospf_configuration()
            else:
                print("Invalid selection")

    ###########################################################################

def snmp_configuration():

    SNMP_file = "C:\Python\XML_Filters\Send_Config\SNMPv2_Send_Config.xml"
    Delete_Config = "C:\Python\XML_Filters\SNMPv2_Delete_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add Configuration")
            print("2. Remove configuration")
            print("3. View SNMP Users")
            print("4. Main Menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                snmp_element = xml.Element("snmp-server")
                native_element.append(snmp_element)
                community_element = xml.Element("community")
                community_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-snmp")
                snmp_element.append(community_element)

                comm_input = input("Please Enter a Community: ")
                comm_id = xml.SubElement(community_element, "name")
                comm_id.text = comm_input

                question_1 = input("Do you want to enter an ACL? ")

                if question_1 == "yes":

                    acl_input = input("Please Enter An Access List: ")
                    acl_element = xml.SubElement(community_element, "access-list-name")
                    acl_element.text = acl_input

                    tree = xml.ElementTree(root) # Writes XML file to share
                    with open(SNMP_file, "wb") as fh:
                        tree.write(fh)

                    view_config_send(SNMP_file) # Call Function

                else:
                    tree = xml.ElementTree(root) # Writes XML file to share
                    with open(SNMP_file, "wb") as fh:
                        tree.write(fh)

                    view_config_send(SNMP_file) # Call Function

            elif config_selection == "2":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                snmp_element = xml.Element("snmp-server")
                native_element.append(snmp_element)
                community_element = xml.Element("community")
                community_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-snmp")
                community_element.set("operation", "delete")
                snmp_element.append(community_element)

                comm_input = input("Please Enter a Community to remove: ")
                comm_id = xml.SubElement(community_element, "name")
                comm_id.set("operation", "delete")
                comm_id.text = comm_input

                tree = xml.ElementTree(root) # Writes XML file to share
                with open(Delete_Config, "wb") as fh:
                    tree.write(fh)

                send_single_configuration(Delete_Config) # Call Function

            elif config_selection == "3":
                view_snmp_users()
            elif config_selection == "4":
                main()
            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. SNMP Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                snmp_configuration()
            else:
                print("Invalid selection")

###########################################################################

def credentials_configuration():

    Delete_User_Config = "C:\Python\XML_Filters\Credentials_Delete_Config.xml"
    Cred_Config = "C:\Python\XML_Filters\Send_Config\Credentials_Send_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add configuration")
            print("2. Remove configuration")
            print("3. View Users")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                username_element = xml.SubElement(native_element, "username")

                print("New User Configuration")

                user_input = input("Please Enter a Username: ")
                user_id = xml.SubElement(username_element, "name")
                user_id.text = user_input

                priv_input = input("Please Enter a Privilage Level: ")
                priv_element = xml.SubElement(username_element, "privilege")
                priv_element.text = priv_input

                password_element = xml.SubElement(username_element, "password")

                password_input = input("Please Enter Password: ")
                pass_element = xml.SubElement(password_element, "password")
                pass_element.text = password_input

                tree = xml.ElementTree(root) # Writes XML file to share
                with open(Cred_Config, "wb") as fh:
                    tree.write(fh)

                view_config_send(Cred_Config) # Call Function

            if config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                username_element = xml.SubElement(native_element, "username")
                username_element.set("xc:operation", "remove")

                question_1 = input("Which user do you want to delete? ")

                name_id = xml.SubElement(username_element, "name")
                name_id.set("xc:operation", "remove")
                name_id.text = question_1

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(Delete_User_Config, "wb") as fh:
                    tree.write(fh)

                view_config_send(Delete_User_Config)  # Call Function

            if config_selection == "3":
                view_users()
            if config_selection == "4":
                main()

            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Credential Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                credentials_configuration()
            else:
                print("Invalid selection")

    ################################################################################

def interface_configuration():

    int_file = "C:\Python\XML_Filters\Send_Config\Interface_Send_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add/Modify configuration")
            print("2. Remove configuration")
            print("3. View Interface(s)")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")
            print("\n")

            if config_selection == "1":

                int_file = "C:\Python\XML_Filters\Send_Config\Interface_Send_Config.xml"
                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                int_element = xml.SubElement(native_element, "interface")

                print("ex. Gigabit Ethernet")
                print("ex. FastEthernet")
                print("ex. Loopback")
                print("ex. Tunnel")
                print("\n")
                int_type = input("Enter an interface type: ")

                while int_type  ==  "Gigabit Ethernet"  or int_type ==  "FastEthernet" or int_type == "Loopback" or int_type == "Tunnel":

                    int_type_leaf = xml.SubElement(int_element, int_type)

                    interface_number = input("Please Enter An Interface number:  ")
                    int_name = xml.SubElement(int_type_leaf, "name")
                    int_name.text = interface_number

                    int_choice_1 = input("Please enter a description: ")
                    int_descrp = xml.SubElement(int_type_leaf, "description")
                    int_descrp.text = int_choice_1

                    ip_leaf = xml.SubElement(int_type_leaf, "ip")
                    address_leaf = xml.SubElement(ip_leaf, "address")
                    primary_leaf = xml.SubElement(address_leaf, "primary")

                    ip_input = input("Please Enter A IP Address: ")
                    address = xml.SubElement(primary_leaf, "address")
                    address.text = ip_input

                    mask_input = input("Please Enter a Subnet Mask: ")
                    mask = xml.SubElement(primary_leaf, "mask")
                    mask.text = mask_input

                    tree = xml.ElementTree(root)
                    with open(int_file, "wb") as fh:
                        tree.write(fh)

                    view_config_send(int_file)

            if config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                int_element = xml.SubElement(native_element, "interface")

                print("ex. GigabitEthernet")
                print("ex. FastEthernet")
                print("ex. Loopback")
                print("ex. Tunnel")
                print("\n")
                int_type = input("Enter an interface type: ")

                while int_type == "GigabitEthernet" or int_type == "FastEthernet" or int_type == "Loopback" or int_type == "Tunnel":

                    int_type = xml.SubElement(int_element, int_type)
                    int_type.set("xc:operation", "remove")

                    interface_number = input("Please Enter An Interface number:  ")
                    int_name = xml.SubElement(int_type, "name")
                    int_name.set("xc:operation", "remove")
                    int_name.text = interface_number

                    tree = xml.ElementTree(root)
                    with open(int_file, "wb") as fh:
                        tree.write(fh)

                    view_config_send(int_file)

            if config_selection == "3":
                view_interfaces()
            elif config_selection == "4":
                main()

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Interface Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                interface_configuration()
            else:
                print("Invalid selection")

################################################################################

def dmvpn_configuration():

    dmvpn_file = "C:\Python\XML_Filters\Send_Config\DMVPN_Send_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add configuration")
            print("2. Remove/Modify configuration")
            print("3. View Interface")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                int_element = xml.SubElement(native_element, "interface")

                ncc_login("device", 830, "cisco", "cisco", {'name': 'csr'})

                int_type_leaf = xml.SubElement(int_element, "Tunnel")

                interface_name = input("Please Enter An Interface: Tunnel: ")
                int_name = xml.SubElement(int_type_leaf, "name")
                int_name.text = interface_name

                int_choice_1 = input("Please enter a description: ")
                int_descrp = xml.SubElement(int_type_leaf, "description")
                int_descrp.text = int_choice_1

                ip_leaf = xml.SubElement(int_type_leaf, "ip")

                address_leaf = xml.SubElement(ip_leaf, "address")
                primary_leaf = xml.SubElement(address_leaf, "primary")

                ip_input = input("Please Enter A IP Address: ")
                address = xml.SubElement(primary_leaf, "address")
                address.text = ip_input

                mask_input = input("Please Enter a Subnet Mask: ")
                mask = xml.SubElement(primary_leaf, "mask")
                mask.text = mask_input

                global nhrp_leaf
                nhrp_leaf = xml.SubElement(ip_leaf, "nhrp")
                nhrp_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-nhrp")

                nhrp_input1 = input("Please Enter NHRP auth ")
                nhrp_auth = xml.SubElement(nhrp_leaf, "authentication")
                nhrp_auth.text = nhrp_input1

                nhrp_input2 = input("Please Enter NHRP group: ")
                nhrp_group = xml.SubElement(nhrp_leaf, "group")
                nhrp_group.text = nhrp_input2

                nhrp_input3 = input("Please Enter NHRP holdtime: ")
                nhrp_hold = xml.SubElement(nhrp_leaf, "holdtime")
                nhrp_hold.text = nhrp_input3

                nhrp_map_leaf = xml.SubElement(nhrp_leaf, "map")
                nhrp_dest_leaf = xml.SubElement(nhrp_map_leaf, "dest-ipv4")

                nhrp_input4 = input("Please Enter Hub Tunnel IP:  ")
                nhrp_hub = xml.SubElement(nhrp_dest_leaf, "dest-ipv4")
                nhrp_hub.text = nhrp_input4

                nhrp_nbma1 = xml.SubElement(nhrp_dest_leaf, "nbma-ipv4")

                nhrp_input4 = input("Please Enter Hub NBMA: ")
                nhrp_hub = xml.SubElement(nhrp_nbma1, "nbma-ipv4")
                nhrp_hub.text = nhrp_input4

                multicast_container = xml.SubElement(nhrp_map_leaf, "multicast")

                nhrp_hub = xml.SubElement(multicast_container, "nbma_ipv4")
                nhrp_hub.text = nhrp_input4

                nhrp_input8 = input("Please Enter NHRP network ID: ")
                nhrp_ID = xml.SubElement(nhrp_leaf, "network-id")
                nhrp_ID.text = nhrp_input8

                nhs_container = xml.SubElement(nhrp_leaf, "nhs")
                global nhs_leaf
                nhs_leaf = xml.SubElement(nhs_container, "ipv4")

                nhrp_input7= input("Please Enter NHRP NHS: ")
                nhrp_nhs = xml.SubElement(nhs_leaf, "ipv4")
                nhrp_nhs.text = nhrp_input7

                tun_leaf = xml.SubElement(int_type_leaf, "tunnel")
                tun_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-tunnel")

                source_input= input("Please Enter a tunnel source: ")
                tun_source = xml.SubElement(tun_leaf, "source")
                tun_source.text = source_input

                ospf_leaf = xml.SubElement(ip_leaf, "ospf")
                ospf_leaf.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-ospf")

                ospf_net_type = input("Please Enter OSPF network Type: ")
                ospf_type = xml.SubElement(ospf_leaf, "network")
                ospf_type.text = ospf_net_type

                tree = xml.ElementTree(root) # Writes XML file to share
                with open(dmvpn_file, "wb") as fh:
                    tree.write(fh)

                view_config_send(dmvpn_file) # Call Function

            if config_selection == "3":
                view_interfaces()
            if config_selection == "4":
                main()
            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. DMVPN Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                dmvpn_configuration()
            else:
                print("Invalid selection")
###################################################################3

def qos_configuration():

    classmap_file = "C:\Python\XML_Filters\Send_Config\QoS_Send_Config.xml"
    policy_map_file = "C:\Python\XML_Filters\Send_Config\PolicyMap_Send_Config.xml"
    serv_policy_file = "C:\Python\XML_Filters\Send_Config\PolicyMap_Shape_Send_Config.xml"
    int_policy_map_file = "C:\Python\XML_Filters\Send_Config\PolicyMap_Interface_Shape_Send_Config.xml"

    config_selection = ' '
    while config_selection != '7':

        try:
            print("\n")
            print("1. Configure Class-map")
            print("2. Configure Policy-map")
            print("3. Confgure Policy-map (shaping)")
            print("4. Apply QoS (interface)")
            print("5. View QoS")
            print("6. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")



            if config_selection == "1":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                policy_element = xml.Element("policy")
                native_element.append(policy_element)
                class_element = xml.Element("class-map")
                class_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
                policy_element.append(class_element)

                print("\n")
                print("Create class-map")
                print("\n")

                global class1_input
                class1_input = input("Please Enter class-map name: ")
                class_id = xml.SubElement(class_element, "name")
                class_id.text = class1_input

                class1_1_input = input("Please Enter class-map type (match-any/all): ")
                prematch_element = xml.SubElement(class_element, "prematch")
                prematch_element.text = class1_1_input

                match_element = xml.SubElement(class_element, "match")

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(classmap_file, "wb") as fh:
                    tree.write(fh)

                print("\n")
                print("1. Configure Match Statement")
                print("2. QoS Configuration Menu")

                config_selection =  input("Please selct an option")
                print("\n")

                if config_selection =="1":

                    while True:

                            print("Please add match statements")
                            print("\n")
                            print("Press CTRL+C to escape at any time")
                            print("\n")

                            DSCP_input = input("Please Enter class-map DSCP/CoS values): ")
                            match_1_element = xml.SubElement(match_element, "dscp")
                            match_1_element.text = DSCP_input

                            tree = xml.ElementTree(root)  # Writes XML file to share
                            with open(classmap_file, "wb") as fh:
                                tree.write(fh)

                if config_selection == "2":
                    qos_configuration()

                else:
                    print("\n")
                    print("Invalid selection")
                    print("\n")

            if config_selection == "2":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                policy_element = xml.Element("policy")
                native_element.append(policy_element)
                policy_map_element = xml.Element("policy-map")
                policy_map_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
                policy_element.append(policy_map_element)

                print("\n")
                print("Policy-map Configuration")
                print("\n")

                config_notice = input("Did you send you last configuration. If not it will be overwritten (yes/no) ").lower()

                if config_notice == "no":
                    view_config_send(policy_map_file)

                elif config_notice == 'yes':

                    pol_map_input = input("Please create policy map name: ")
                    pol_map_name = xml.SubElement(policy_map_element, "name")
                    pol_map_name.text = pol_map_input

                    pol_class1_element = xml.SubElement(policy_map_element, "class")

                    while config_notice == 'yes':

                        print("\n")
                        print("Press CTRL+C to escape at any time")
                        print("\n")

                        class_to_queue = input("Please map class to queues: ")
                        pol_name_1 = xml.SubElement(pol_class1_element, "name")
                        pol_name_1.text = class_to_queue

                        action_list_element = xml.SubElement(pol_class1_element, "action-list")


                        print("\n")
                        print("1. Bandwidth")
                        print("2. Priority")
                        print("\n")

                        selection= " "
                        while selection != "1" or "2":

                            selection = input("Please select an option: ")

                            if selection == "1" or "2":
                                break

                        if selection == "1":

                            action_type_element = xml.SubElement(action_list_element, "action-type")
                            action_type_element.text = "bandwidth"

                            priority_container = xml.SubElement(action_list_element, "bandwidth")

                            bandwidth_input = input("Please enter bandwith percent")
                            percent_element = xml.SubElement(priority_container, "percent")
                            percent_element.text = bandwidth_input

                            tree = xml.ElementTree(root)  # Writes XML file to share
                            with open(policy_map_file, "wb") as fh:
                                tree.write(fh)

                        if selection == "2":

                            action_type_element = xml.SubElement(action_list_element, "action-type")
                            action_type_element.text = "priority"

                            priority_container = xml.SubElement(action_list_element, "priority")

                            bandwidth_input = input("Please enter priority percent")
                            percent_element = xml.SubElement(priority_container, "percent")
                            percent_element.text = bandwidth_input

                            tree = xml.ElementTree(root)  # Writes XML file to share
                            with open(policy_map_file, "wb") as fh:
                                tree.write(fh)

            if config_selection == "3":

                root = xml.Element("config")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                policy_element = xml.Element("policy")
                native_element.append(policy_element)
                policy_map_element = xml.Element("policy-map")
                policy_map_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")
                policy_element.append(policy_map_element)

                pol_map_5_input = input("Please create policy map name: ")
                pol_map_5_name = xml.SubElement(policy_map_element, "name")
                pol_map_5_name.text = pol_map_5_input

                pol_class5_element = xml.SubElement(policy_map_element, "class")

                nest_input = input("Please enter a nested class (class-default): ")
                nested_name = xml.SubElement(pol_class5_element, "name")
                nested_name.text = nest_input

                action_list_5_element = xml.SubElement(pol_class5_element, "action-list")

                while True:

                    print("\n")
                    print("1. Shape")
                    print("2. Police")
                    print("\n")

                    selection = input("Please select an option: ")

                    if selection == "1" or "2":

                        if selection == "1":

                            action_type_5_element = xml.SubElement(action_list_5_element, "action-type")
                            action_type_5_element.text = "shape"

                            shape_element = xml.SubElement(action_list_5_element, "shape")
                            average_element = xml.SubElement(shape_element, "average")

                            bandwidth_5_input = input("Please enter bandwith in bits: ")
                            percent_5_element = xml.SubElement(average_element, "bit-rate")
                            percent_5_element.text = bandwidth_5_input

                            action_list_element = xml.SubElement(pol_class5_element, "action-list")
                            action_type_element = xml.SubElement(action_list_element, "action-type")
                            action_type_element.text = "service-policy"

                            parent_pol_input = input("Please enter a child policy map: ")
                            action_list_5 = xml.SubElement(action_list_element, "service-policy")
                            action_list_5.text = parent_pol_input

                            tree = xml.ElementTree(root)
                            with open(serv_policy_file, "wb") as fh:
                                tree.write(fh)

                            view_config_send(serv_policy_file)

                        elif selection == "2":

                            action_type_5_element = xml.SubElement(action_list_5_element, "action-type")
                            action_type_5_element.text = "police"

                            tar_bit_rate = xml.SubElement(action_list_5_element, "police-target-bitrate")
                            police = xml.SubElement(tar_bit_rate, "police")

                            bandwidth_5_input = input("Please enter bandwith in bits: ")
                            percent_5_element = xml.SubElement(police, "bit-rate")
                            percent_5_element.text = bandwidth_5_input

                            action_list_element = xml.SubElement(pol_class5_element, "action-list")
                            action_type_element = xml.SubElement(action_list_element, "action-type")
                            action_type_element.text = "service-policy"

                            parent_pol_input = input("Please enter a child policy map: ")
                            action_list_5 = xml.SubElement(action_list_element, "service-policy")
                            action_list_5.text = parent_pol_input

                            tree = xml.ElementTree(root)
                            with open(serv_policy_file, "wb") as fh:
                                tree.write(fh)

                            view_config_send(serv_policy_file)

                        else:
                            print("Invalid input")


            if config_selection == "4":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)

                int_element = xml.SubElement(native_element, "interface")

                int_type = input("Please enter interface type: ")
                int_type_element= xml.SubElement(int_element, int_type)

                int_num = input("Please enter interface number:")
                int_num_element = xml.SubElement(int_type_element, "name")
                int_num_element.text = int_num

                service_1_element = xml.SubElement(int_type_element, "service-policy")
                service_1_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-policy")

                output_1_policy = input("Please enter service policy: ")
                service_1_output = xml.SubElement(service_1_element, "output")
                service_1_output.text = output_1_policy

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(int_policy_map_file, "wb") as fh:
                    tree.write(fh)

                view_config_send(int_policy_map_file)  # Call Function

            if config_selection == "5":
                view_qos()
            if config_selection == "6":
                main()
            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Send Class-map Configuration")
            print("3. Send Policy Configuration")
            print("4. QoS Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                view_config_send(classmap_file)
            if escape_1 == "3":
                view_config_send(policy_map_file)
            if escape_1 == "4":
                qos_configuration()
            else:
                print("Invalid selection")

def tacacs_configuration():

    config_selection = ' '
    while config_selection != '4':

        tacacs_file = "C:\Python\XML_Filters\Send_Config\TACACS_Send_Config.xml"
        del_tacacs = "C:\Python\XML_Filters\Send_Config\TACACS_Delete_Config.xml"

        try:

            print("\n")
            print("1. Add configuration")
            print("2. Remove/Modify configuration")
            print("3. View TACACS")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)
                tacacs_elem = xml.SubElement(native_element, "tacacs")
                tac_server = xml.SubElement(tacacs_elem, "server")
                tac_server.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-aaa")

                print("\n")
                print("Configure TACACS Server:")
                print("\n")

                tac_name_input = input("Please enter a TACACS group: ")
                tac_name = xml.SubElement(tac_server, "name")
                tac_name.text = tac_name_input

                tac_address = xml.SubElement(tac_server, "address")

                tac_ipv4_input = input("Please enter a TACACS server IP: ")
                tac_ipv4 = xml.SubElement(tac_address, "ipv4")
                tac_ipv4.text = tac_ipv4_input

                tac_key_elem = xml.SubElement(tac_server, "key")

                tac_key_input = input("Please enter a TACACS key: ")
                tac_key = xml.SubElement(tac_key_elem, "key")
                tac_key.text = tac_key_input

                tree = xml.ElementTree(root)  # Writes XML file to share
                with open(tacacs_file, "wb") as fh:
                    tree.write(fh)

                view_config_send(tacacs_file)

            elif config_selection == "2":

                print("\n")
                print("Configure TACACS Server:")
                print("\n")

                print("\n")
                print("1. Remove TACACS Group")
                print("2. Remove/Modify TACACS Group")
                print("\n")
                print("Press CTRL+C to escape at any time")
                print("\n")

                config_selection = input("Please select an option:  ")

                if config_selection == "1":

                    del_tacacs = "C:\Python\XML_Filters\Send_Config\TACACS_Delete_Config.xml"
                    root = xml.Element("config")
                    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    native_element = xml.Element("native")
                    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                    root.append(native_element)

                    tacacs_elem = xml.SubElement(native_element, "tacacs")
                    tac_server = xml.SubElement(tacacs_elem, "server")
                    tac_server.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-aaa")

                    tac_name_input = input("Please enter a TACACS group: ")
                    tac_name = xml.SubElement(tac_server, "name")
                    tac_name.set("xc:operation", "remove")
                    tac_name.text = tac_name_input

                    view_config_send(del_tacacs)

                if config_selection == "2":

                    root = xml.Element("config")
                    root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                    native_element = xml.Element("native")
                    native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                    root.append(native_element)

                    tacacs_elem = xml.SubElement(native_element, "tacacs")
                    tac_server = xml.SubElement(tacacs_elem, "server")
                    tac_server.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-aaa")

                    print("\n")
                    print("1. Add/Remove TACACS Server")
                    print("\n")
                    print("Press CTRL+C to escape at any time")
                    print("\n")

                    config_selection = input("Please select an option:  ")

                    if config_selection == "1":

                        tac_name_input = input("Please enter a TACACS group: ")
                        tac_name = xml.SubElement(tac_server, "name")
                        tac_name.text = tac_name_input

                        tac_address = xml.SubElement(tac_server, "address")

                        tac_ipv4_input = input("Please enter a TACACS server IP: ")
                        tac_ipv4 = xml.SubElement(tac_address, "ipv4")
                        tac_ipv4.set("xc:operation", "remove")
                        tac_ipv4.text = tac_ipv4_input

                        tree = xml.ElementTree(root)  # Writes XML file to share
                        with open(del_tacacs, "wb") as fh:
                            tree.write(fh)

                        view_config_send(del_tacacs)

            elif config_selection == "3":
                view_tacacs()
            elif config_selection == "4":
                main()
            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. TACACS Configuration Menu")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                tacacs_configuration()
            else:
                print("Invalid selection")

def prefix_configuration():

    prefix_file = "C:\Python\XML_Filters\Send_Config\Prefix_Send_Config.xml"
    prefix_del = "C:\Python\XML_Filters\Send_Config\Prefix_Delete_Config.xml"

    config_selection = ' '
    while config_selection != '4':

        try:

            print("\n")
            print("1. Add configuration")
            print("2. Remove/Modify configuration")
            print("3. View Prefix-list")
            print("4. Main menu")
            print("\n")
            print("Press CTRL+C to escape at any time")
            print("\n")

            config_selection = input("Please select an option:  ")

            if config_selection == "1":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)

                print("\n")
                print("Configure Prefix-List:")
                print("\n")

                prefix = xml.SubElement(native_element, "ip")
                prefix_list = xml.SubElement(prefix, "prefix-list")
                prefixes = xml.SubElement(prefix_list, "prefixes")

                prefix_name_input = input("Please enter a prefix-list name: ")
                prefix_name = xml.SubElement(prefixes, "name")
                prefix_name.text = prefix_name_input


                while config_selection == "1":

                    seq = xml.SubElement(prefixes, "seq")

                    seq_input = input("Please enter a seq number: ")
                    num = xml.SubElement(seq, "no")
                    num.text = seq_input

                    permit_deny = xml.SubElement(seq, "permit")

                    prefix_input = input("Please enter a prefix: ")
                    ip = xml.SubElement(permit_deny, "ip")
                    ip.text =prefix_input

                    tree = xml.ElementTree(root)  # Writes XML file to share
                    with open(prefix_file, "wb") as fh:
                        tree.write(fh)

            elif config_selection == "2":

                root = xml.Element("config")
                root.set("xmlns", "urn:ietf:params:xml:ns:netconf:base:1.0")
                root.set("xmlns:xc", "urn:ietf:params:xml:ns:netconf:base:1.0")
                native_element = xml.Element("native")
                native_element.set("xmlns", "http://cisco.com/ns/yang/Cisco-IOS-XE-native")
                root.append(native_element)

                print("\n")
                print("Modify/Delete  Prefix-List:")
                print("\n")

                prefix = xml.SubElement(native_element, "ip")
                prefix_list = xml.SubElement(prefix, "prefix-list")
                prefixes = xml.SubElement(prefix_list, "prefixes")

                print("\n")
                print("1. Remove Prefix List")
                print("2. Remove/Modify Prefix Statement")
                print("\n")
                print("Press CTRL+C to escape at any time")
                print("\n")

                config_selection = input("Please select an option:  ")

                if config_selection == "1":

                    prefix_name_input = input("Please enter a prefix-list name: ")
                    prefix_name = xml.SubElement(prefixes, "name")
                    prefix_name.text = prefix_name_input

                    tree = xml.ElementTree(root)  # Writes XML file to share
                    with open(prefix_del, "wb") as fh:
                        tree.write(fh)

                    view_config_send(prefix_del)

                if config_selection == "2":

                    prefix_name_input = input("Please enter a prefix-list name: ")
                    prefix_name = xml.SubElement(prefixes, "name")
                    prefix_name.text = prefix_name_input

                    seq = xml.SubElement(prefixes, "seq")
                    seq.set("xc:operation", "remove")

                    seq_input = input("Please enter a seq number: ")
                    num = xml.SubElement(seq, "no")
                    num.set("xc:operation", "remove")
                    num.text = seq_input

                    tree = xml.ElementTree(root)  # Writes XML file to share
                    with open(prefix_del, "wb") as fh:
                        tree.write(fh)

                    view_config_send(prefix_del)

            elif config_selection == "3":
                view_prefix_list()

            elif config_selection == "4":
                main()

            else:
                print("\n")
                print("Invalid selection")
                print("\n")

        except KeyboardInterrupt:

            print("\n")
            print("1. Main Menu")
            print("2. Prefix-List Configuration Menu")
            print("3. Send Prefix- List Configuration ")

            escape_1 = input("What menu do you want to escape to?")

            if escape_1 == "1":
                main()
            if escape_1 == "2":
                prefix_configuration()
            if escape_1 == "3":
                view_config_send(prefix_file)
            else:
                print("Invalid selection")


if __name__ == '__main__':

    main()
